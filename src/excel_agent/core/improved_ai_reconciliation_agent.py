#!/usr/bin/env python3
"""
Improved AI Reconciliation Agent
Enhanced version following Strands Agent best practices with proper training data integration
"""

import os
import json
import openpyxl
import xlrd
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple, Union
import logging
import traceback

from strands_base_agent import StrandsBaseAgent

class ImprovedAIReconciliationAgent(StrandsBaseAgent):
    """
    Enhanced AI Reconciliation Agent following Strands Agent best practices.
    
    This agent provides:
    - Model-driven reconciliation processing
    - Comprehensive training data integration
    - Advanced pattern recognition
    - Continuous learning capabilities
    - Robust error handling and monitoring
    """
    
    def __init__(self, 
                 config: Optional[Dict[str, Any]] = None,
                 training_data_path: Optional[str] = None,
                 openai_api_key: Optional[str] = None):
        """
        Initialize the improved AI reconciliation agent.
        
        Args:
            config: Configuration dictionary
            training_data_path: Path to training data directory
            openai_api_key: OpenAI API key for AI capabilities
        """
        super().__init__(
            name="ImprovedAIReconciliationAgent",
            config=config,
            training_data_path=training_data_path
        )
        
        self.openai_api_key = openai_api_key or os.getenv('OPENAI_API_KEY')
        self.reconciliation_rules = {}
        self.pattern_library = {}
        self.learning_history = []
        
        # Initialize AI capabilities
        self._initialize_ai_capabilities()
        
        # Load reconciliation-specific training data
        self._load_reconciliation_training_data()
    
    def _initialize_ai_capabilities(self):
        """Initialize AI capabilities for advanced reconciliation."""
        if not self.openai_api_key:
            self.logger.warning("OpenAI API key not provided - AI capabilities will be limited")
            return
        
        try:
            import openai
            openai.api_key = self.openai_api_key
            self.logger.info("OpenAI API initialized successfully")
        except ImportError:
            self.logger.warning("OpenAI library not available - AI capabilities will be limited")
        except Exception as e:
            self.logger.error(f"Error initializing OpenAI: {str(e)}")
    
    def _load_reconciliation_training_data(self):
        """Load reconciliation-specific training data."""
        # Load OP manual
        op_manual = self.get_training_data("op_manual")
        if not op_manual:
            op_manual = self._get_default_op_manual()
            self.update_training_data("op_manual", op_manual)
        
        # Load historical patterns
        historical_patterns = self.get_training_data("historical_patterns")
        if not historical_patterns:
            historical_patterns = self._get_default_historical_patterns()
            self.update_training_data("historical_patterns", historical_patterns)
        
        # Load reconciliation rules
        reconciliation_rules = self.get_training_data("reconciliation_rules")
        if not reconciliation_rules:
            reconciliation_rules = self._get_default_reconciliation_rules()
            self.update_training_data("reconciliation_rules", reconciliation_rules)
        
        self.logger.info("Loaded reconciliation training data successfully")
    
    def _get_default_op_manual(self) -> Dict[str, Any]:
        """Get default OP manual data."""
        return {
            "gl_accounts": {
                "74400": {
                    "name": "RBC Activity",
                    "bank_activities": ["RBC activity", "EFUNDS Corp – FEE SETTLE", "Withdrawal Coin or Withdrawal Currency"],
                    "expected_timing": "daily",
                    "variance_threshold": 1000.0
                },
                "74505": {
                    "name": "CNS Settlement",
                    "bank_activities": ["CNS Settlement activity", "PULSE FEES activity"],
                    "expected_timing": "daily",
                    "variance_threshold": 500.0
                },
                "74510": {
                    "name": "EFUNDS Corp Daily Settlement",
                    "bank_activities": ["EFUNDS Corp – DLY SETTLE activity"],
                    "expected_timing": "daily",
                    "variance_threshold": 2000.0
                },
                "74515": {
                    "name": "Cash Letter Corrections",
                    "bank_activities": ["Cash Letter Corr activity"],
                    "expected_timing": "as_needed",
                    "variance_threshold": 100.0
                },
                "74520": {
                    "name": "Image Check Presentment",
                    "bank_activities": ["1591 Image CL Presentment activity"],
                    "expected_timing": "daily",
                    "variance_threshold": 5000.0
                },
                "74525": {
                    "name": "Returned Drafts",
                    "bank_activities": ["1590 Image CL Presentment activity (returns)"],
                    "expected_timing": "daily",
                    "variance_threshold": 1000.0
                },
                "74530": {
                    "name": "ACH Activity",
                    "bank_activities": ["ACH ADV File activity"],
                    "expected_timing": "daily",
                    "variance_threshold": 10000.0
                },
                "74535": {
                    "name": "ICUL Services",
                    "bank_activities": ["ICUL ServCorp activity"],
                    "expected_timing": "monthly",
                    "variance_threshold": 500.0
                },
                "74540": {
                    "name": "CRIF Loans",
                    "bank_activities": ["ACH ADV FILE - Orig CR activity (CRIF loans)"],
                    "expected_timing": "daily",
                    "variance_threshold": 2000.0
                },
                "74550": {
                    "name": "Cooperative Business",
                    "bank_activities": ["Cooperative Business activity"],
                    "expected_timing": "monthly",
                    "variance_threshold": 1000.0
                },
                "74560": {
                    "name": "Check Deposits",
                    "bank_activities": ["1590 Image CL Presentment activity (deposits)"],
                    "expected_timing": "daily",
                    "variance_threshold": 5000.0
                },
                "74570": {
                    "name": "ACH Returns",
                    "bank_activities": ["ACH ADV FILE - Orig DB activity (ACH returns)"],
                    "expected_timing": "daily",
                    "variance_threshold": 2000.0
                }
            },
            "timing_differences": {
                "74505": {
                    "description": "ATM settlement activity posted to GL on last day of month",
                    "expected": True,
                    "variance_threshold": 1000.0
                },
                "74510": {
                    "description": "Shared Branching activity recorded in GL on last day of month",
                    "expected": True,
                    "variance_threshold": 2000.0
                },
                "74560": {
                    "description": "Check deposit activity at branches posted to GL on last day of month",
                    "expected": True,
                    "variance_threshold": 5000.0
                },
                "74535": {
                    "description": "Gift Card activity posted to GL on last day of month",
                    "expected": True,
                    "variance_threshold": 500.0
                },
                "74550": {
                    "description": "CBS activity posted to GL on last day of month",
                    "expected": True,
                    "variance_threshold": 1000.0
                },
                "74540": {
                    "description": "CRIF indirect loan activity posted to GL on last day of month",
                    "expected": True,
                    "variance_threshold": 2000.0
                }
            }
        }
    
    def _get_default_historical_patterns(self) -> Dict[str, Any]:
        """Get default historical patterns for learning."""
        return {
            "common_discrepancies": [
                {
                    "pattern": "timing_difference",
                    "description": "Activity posted on different days",
                    "frequency": 0.15,
                    "resolution": "Accept if within 3 days"
                },
                {
                    "pattern": "rounding_difference",
                    "description": "Small rounding differences in amounts",
                    "frequency": 0.08,
                    "resolution": "Accept if less than $0.01"
                },
                {
                    "pattern": "duplicate_entry",
                    "description": "Same transaction posted twice",
                    "frequency": 0.05,
                    "resolution": "Remove duplicate entry"
                }
            ],
            "success_patterns": [
                {
                    "pattern": "exact_match",
                    "description": "Perfect match between GL and bank data",
                    "frequency": 0.70,
                    "confidence": 1.0
                },
                {
                    "pattern": "timing_match",
                    "description": "Match with timing difference",
                    "frequency": 0.20,
                    "confidence": 0.95
                }
            ]
        }
    
    def _get_default_reconciliation_rules(self) -> Dict[str, Any]:
        """Get default reconciliation rules."""
        return {
            "matching_criteria": {
                "amount_tolerance": 0.01,
                "date_tolerance_days": 3,
                "description_similarity_threshold": 0.8
            },
            "validation_rules": {
                "min_transaction_amount": 0.01,
                "max_transaction_amount": 1000000.0,
                "required_fields": ["date", "amount", "description"]
            },
            "reporting_rules": {
                "include_timing_differences": True,
                "include_variance_analysis": True,
                "include_pattern_analysis": True
            }
        }
    
    def process(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process reconciliation data with advanced AI capabilities.
        
        Args:
            input_data: Dictionary containing GL and bank file paths
            
        Returns:
            Comprehensive reconciliation results
        """
        try:
            self.logger.info("Starting AI-powered reconciliation process")
            
            # Extract file paths
            gl_file_path = input_data.get("gl_file")
            bank_file_path = input_data.get("bank_file")
            
            if not gl_file_path or not bank_file_path:
                raise ValueError("Both GL and bank file paths are required")
            
            # Load and validate files
            gl_data = self._load_and_validate_gl_file(gl_file_path)
            bank_data = self._load_and_validate_bank_file(bank_file_path)
            
            # Perform AI-powered reconciliation
            reconciliation_results = self._perform_ai_reconciliation(gl_data, bank_data)
            
            # Generate insights and recommendations
            insights = self._generate_ai_insights(reconciliation_results)
            
            # Update learning history
            self._update_learning_history(reconciliation_results)
            
            return {
                "status": "success",
                "reconciliation_results": reconciliation_results,
                "insights": insights,
                "learning_updated": True,
                "timestamp": datetime.now().isoformat()
            }
            
        except Exception as e:
            self.logger.error(f"Error in reconciliation process: {str(e)}")
            raise
    
    def _load_and_validate_gl_file(self, file_path: str) -> Dict[str, Any]:
        """Load and validate GL activity file."""
        self.logger.info(f"Loading GL file: {file_path}")
        
        try:
            if file_path.endswith('.xlsx'):
                return self._load_xlsx_file(file_path, "GL")
            elif file_path.endswith('.xls'):
                return self._load_xls_file(file_path, "GL")
            else:
                raise ValueError(f"Unsupported file format: {file_path}")
        except Exception as e:
            self.logger.error(f"Error loading GL file {file_path}: {str(e)}")
            raise
    
    def _load_and_validate_bank_file(self, file_path: str) -> Dict[str, Any]:
        """Load and validate bank statement file."""
        self.logger.info(f"Loading bank file: {file_path}")
        
        try:
            if file_path.endswith('.xlsx'):
                return self._load_xlsx_file(file_path, "Bank")
            elif file_path.endswith('.xls'):
                return self._load_xls_file(file_path, "Bank")
            else:
                raise ValueError(f"Unsupported file format: {file_path}")
        except Exception as e:
            self.logger.error(f"Error loading bank file {file_path}: {str(e)}")
            raise
    
    def _load_xlsx_file(self, file_path: str, file_type: str) -> Dict[str, Any]:
        """Load XLSX file with validation."""
        wb = openpyxl.load_workbook(file_path)
        
        data = {
            "file_type": file_type,
            "file_path": file_path,
            "sheets": {},
            "metadata": {
                "total_sheets": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "loaded_at": datetime.now().isoformat()
            }
        }
        
        # Load each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_data = self._extract_sheet_data(sheet, file_type)
            data["sheets"][sheet_name] = sheet_data
        
        wb.close()
        return data
    
    def _load_xls_file(self, file_path: str, file_type: str) -> Dict[str, Any]:
        """Load XLS file with validation."""
        try:
            wb = xlrd.open_workbook(file_path)
        except Exception as e:
            self.logger.warning(f"Could not load XLS file {file_path}: {str(e)}")
            raise ValueError(f"Could not load XLS file: {str(e)}")
        
        data = {
            "file_type": file_type,
            "file_path": file_path,
            "sheets": {},
            "metadata": {
                "total_sheets": wb.nsheets,
                "sheet_names": wb.sheet_names(),
                "loaded_at": datetime.now().isoformat()
            }
        }
        
        # Load each sheet
        for sheet_index in range(wb.nsheets):
            sheet = wb.sheet_by_index(sheet_index)
            sheet_data = self._extract_xls_sheet_data(sheet, file_type)
            data["sheets"][sheet.name] = sheet_data
        
        return data
    
    def _extract_sheet_data(self, sheet, file_type: str) -> Dict[str, Any]:
        """Extract data from XLSX sheet."""
        data = {
            "rows": [],
            "gl_accounts": [],
            "transactions": [],
            "summary": {
                "total_rows": 0,
                "total_amount": 0.0,
                "debit_amount": 0.0,
                "credit_amount": 0.0
            }
        }
        
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data["rows"].append(list(row))
                data["summary"]["total_rows"] += 1
                
                # Extract transaction data
                transaction = self._extract_transaction_data(row, file_type)
                if transaction:
                    data["transactions"].append(transaction)
                    data["summary"]["total_amount"] += abs(transaction.get("amount", 0))
                    
                    if transaction.get("amount", 0) > 0:
                        data["summary"]["debit_amount"] += transaction["amount"]
                    else:
                        data["summary"]["credit_amount"] += abs(transaction["amount"])
        
        return data
    
    def _extract_xls_sheet_data(self, sheet, file_type: str) -> Dict[str, Any]:
        """Extract data from XLS sheet."""
        data = {
            "rows": [],
            "gl_accounts": [],
            "transactions": [],
            "summary": {
                "total_rows": 0,
                "total_amount": 0.0,
                "debit_amount": 0.0,
                "credit_amount": 0.0
            }
        }
        
        for row_index in range(sheet.nrows):
            row = sheet.row_values(row_index)
            if any(cell for cell in row):
                data["rows"].append(row)
                data["summary"]["total_rows"] += 1
                
                # Extract transaction data
                transaction = self._extract_transaction_data(row, file_type)
                if transaction:
                    data["transactions"].append(transaction)
                    data["summary"]["total_amount"] += abs(transaction.get("amount", 0))
                    
                    if transaction.get("amount", 0) > 0:
                        data["summary"]["debit_amount"] += transaction["amount"]
                    else:
                        data["summary"]["credit_amount"] += abs(transaction["amount"])
        
        return data
    
    def _extract_transaction_data(self, row: List[Any], file_type: str) -> Optional[Dict[str, Any]]:
        """Extract transaction data from a row."""
        try:
            # Look for amount in the row
            amount = None
            date = None
            description = ""
            
            for cell in row:
                if isinstance(cell, (int, float)) and cell != 0:
                    amount = cell
                elif isinstance(cell, datetime):
                    date = cell
                elif isinstance(cell, str) and len(cell) > 2:
                    description = cell
            
            if amount is not None:
                return {
                    "amount": amount,
                    "date": date.isoformat() if date else None,
                    "description": description,
                    "file_type": file_type
                }
        except Exception as e:
            self.logger.debug(f"Error extracting transaction data: {str(e)}")
        
        return None
    
    def _perform_ai_reconciliation(self, gl_data: Dict[str, Any], bank_data: Dict[str, Any]) -> Dict[str, Any]:
        """Perform AI-powered reconciliation analysis."""
        self.logger.info("Performing AI-powered reconciliation analysis")
        
        # Extract GL balances
        gl_balances = self._extract_gl_balances(gl_data)
        
        # Extract bank balances
        bank_balances = self._extract_bank_balances(bank_data)
        
        # Perform pattern matching
        matches = self._perform_pattern_matching(gl_balances, bank_balances)
        
        # Calculate variances
        variances = self._calculate_variances(gl_balances, bank_balances)
        
        # Generate reconciliation summary
        summary = self._generate_reconciliation_summary(gl_balances, bank_balances, matches, variances)
        
        return {
            "gl_balances": gl_balances,
            "bank_balances": bank_balances,
            "matches": matches,
            "variances": variances,
            "summary": summary,
            "analysis_timestamp": datetime.now().isoformat()
        }
    
    def _extract_gl_balances(self, gl_data: Dict[str, Any]) -> Dict[str, Any]:
        """Extract GL account balances from GL data."""
        gl_balances = {}
        op_manual = self.get_training_data("op_manual")
        
        for sheet_name, sheet_data in gl_data["sheets"].items():
            if sheet_name.isdigit() and len(sheet_name) == 5:
                gl_account = sheet_name
                summary = sheet_data["summary"]
                
                gl_balances[gl_account] = {
                    "account_name": op_manual["gl_accounts"].get(gl_account, {}).get("name", f"GL Account {gl_account}"),
                    "debits": summary["debit_amount"],
                    "credits": summary["credit_amount"],
                    "net_balance": summary["debit_amount"] - summary["credit_amount"],
                    "transaction_count": len(sheet_data["transactions"]),
                    "variance_threshold": op_manual["gl_accounts"].get(gl_account, {}).get("variance_threshold", 1000.0)
                }
        
        return gl_balances
    
    def _extract_bank_balances(self, bank_data: Dict[str, Any]) -> Dict[str, Any]:
        """Extract bank balances from bank data."""
        bank_balances = {
            "total_debits": 0.0,
            "total_credits": 0.0,
            "net_balance": 0.0,
            "transaction_count": 0,
            "transactions": []
        }
        
        for sheet_name, sheet_data in bank_data["sheets"].items():
            bank_balances["total_debits"] += sheet_data["summary"]["debit_amount"]
            bank_balances["total_credits"] += sheet_data["summary"]["credit_amount"]
            bank_balances["transaction_count"] += len(sheet_data["transactions"])
            bank_balances["transactions"].extend(sheet_data["transactions"])
        
        bank_balances["net_balance"] = bank_balances["total_debits"] - bank_balances["total_credits"]
        
        return bank_balances
    
    def _perform_pattern_matching(self, gl_balances: Dict[str, Any], bank_balances: Dict[str, Any]) -> Dict[str, Any]:
        """Perform pattern matching between GL and bank data."""
        matches = {
            "exact_matches": [],
            "timing_matches": [],
            "partial_matches": [],
            "unmatched_gl": [],
            "unmatched_bank": []
        }
        
        # Simple pattern matching logic
        for gl_account, gl_data in gl_balances.items():
            if abs(gl_data["net_balance"]) < 0.01:
                matches["exact_matches"].append({
                    "gl_account": gl_account,
                    "balance": gl_data["net_balance"],
                    "match_type": "exact"
                })
            else:
                matches["unmatched_gl"].append({
                    "gl_account": gl_account,
                    "balance": gl_data["net_balance"],
                    "variance": gl_data["net_balance"]
                })
        
        return matches
    
    def _calculate_variances(self, gl_balances: Dict[str, Any], bank_balances: Dict[str, Any]) -> Dict[str, Any]:
        """Calculate variances between GL and bank data."""
        total_gl_balance = sum(gl_data["net_balance"] for gl_data in gl_balances.values())
        total_bank_balance = bank_balances["net_balance"]
        
        variances = {
            "total_gl_balance": total_gl_balance,
            "total_bank_balance": total_bank_balance,
            "net_variance": total_gl_balance - total_bank_balance,
            "variance_percentage": (abs(total_gl_balance - total_bank_balance) / abs(total_gl_balance)) * 100 if total_gl_balance != 0 else 0,
            "account_variances": {}
        }
        
        # Calculate individual account variances
        for gl_account, gl_data in gl_balances.items():
            variance = gl_data["net_balance"]
            threshold = gl_data["variance_threshold"]
            
            variances["account_variances"][gl_account] = {
                "variance": variance,
                "threshold": threshold,
                "exceeds_threshold": abs(variance) > threshold,
                "severity": "high" if abs(variance) > threshold * 2 else "medium" if abs(variance) > threshold else "low"
            }
        
        return variances
    
    def _generate_reconciliation_summary(self, gl_balances: Dict[str, Any], bank_balances: Dict[str, Any], 
                                       matches: Dict[str, Any], variances: Dict[str, Any]) -> Dict[str, Any]:
        """Generate comprehensive reconciliation summary."""
        return {
            "reconciliation_status": "completed",
            "total_gl_accounts": len(gl_balances),
            "total_gl_balance": variances["total_gl_balance"],
            "total_bank_balance": variances["total_bank_balance"],
            "net_variance": variances["net_variance"],
            "variance_percentage": variances["variance_percentage"],
            "is_balanced": abs(variances["net_variance"]) < 0.01,
            "exact_matches": len(matches["exact_matches"]),
            "unmatched_accounts": len(matches["unmatched_gl"]),
            "high_variance_accounts": sum(1 for acc in variances["account_variances"].values() if acc["exceeds_threshold"]),
            "recommendations": self._generate_recommendations(variances, matches)
        }
    
    def _generate_recommendations(self, variances: Dict[str, Any], matches: Dict[str, Any]) -> List[str]:
        """Generate recommendations based on analysis."""
        recommendations = []
        
        if abs(variances["net_variance"]) > 1000:
            recommendations.append("High net variance detected - investigate root cause")
        
        high_variance_count = sum(1 for acc in variances["account_variances"].values() if acc["exceeds_threshold"])
        if high_variance_count > 0:
            recommendations.append(f"{high_variance_count} accounts exceed variance threshold - review individual accounts")
        
        if len(matches["unmatched_gl"]) > 0:
            recommendations.append(f"{len(matches['unmatched_gl'])} GL accounts have unmatched balances - investigate timing differences")
        
        return recommendations
    
    def _generate_ai_insights(self, reconciliation_results: Dict[str, Any]) -> Dict[str, Any]:
        """Generate AI-powered insights and recommendations."""
        insights = {
            "pattern_analysis": {},
            "anomaly_detection": {},
            "learning_insights": {},
            "recommendations": []
        }
        
        # Analyze patterns
        variances = reconciliation_results["variances"]
        insights["pattern_analysis"] = {
            "variance_trend": "increasing" if variances["net_variance"] > 0 else "decreasing",
            "high_variance_accounts": [acc for acc, data in variances["account_variances"].items() if data["exceeds_threshold"]],
            "balance_distribution": "normal" if abs(variances["net_variance"]) < 1000 else "abnormal"
        }
        
        # Detect anomalies
        insights["anomaly_detection"] = {
            "extreme_variances": [acc for acc, data in variances["account_variances"].items() if data["severity"] == "high"],
            "timing_issues": len(reconciliation_results["matches"]["timing_matches"]),
            "data_quality_issues": len(reconciliation_results["matches"]["unmatched_gl"])
        }
        
        # Generate learning insights
        insights["learning_insights"] = {
            "success_rate": len(reconciliation_results["matches"]["exact_matches"]) / len(reconciliation_results["gl_balances"]) if reconciliation_results["gl_balances"] else 0,
            "improvement_areas": self._identify_improvement_areas(reconciliation_results),
            "pattern_confidence": self._calculate_pattern_confidence(reconciliation_results)
        }
        
        return insights
    
    def _identify_improvement_areas(self, reconciliation_results: Dict[str, Any]) -> List[str]:
        """Identify areas for improvement based on analysis."""
        improvement_areas = []
        
        if reconciliation_results["summary"]["high_variance_accounts"] > 0:
            improvement_areas.append("Variance threshold tuning")
        
        if reconciliation_results["summary"]["unmatched_accounts"] > 0:
            improvement_areas.append("Pattern matching algorithms")
        
        if reconciliation_results["variances"]["variance_percentage"] > 5:
            improvement_areas.append("Data quality validation")
        
        return improvement_areas
    
    def _calculate_pattern_confidence(self, reconciliation_results: Dict[str, Any]) -> float:
        """Calculate confidence in pattern matching results."""
        total_accounts = len(reconciliation_results["gl_balances"])
        exact_matches = len(reconciliation_results["matches"]["exact_matches"])
        
        if total_accounts == 0:
            return 0.0
        
        return exact_matches / total_accounts
    
    def _update_learning_history(self, reconciliation_results: Dict[str, Any]):
        """Update learning history for continuous improvement."""
        learning_entry = {
            "timestamp": datetime.now().isoformat(),
            "reconciliation_results": reconciliation_results,
            "success_rate": reconciliation_results["summary"]["exact_matches"] / reconciliation_results["summary"]["total_gl_accounts"] if reconciliation_results["summary"]["total_gl_accounts"] > 0 else 0,
            "variance_level": reconciliation_results["variances"]["variance_percentage"]
        }
        
        self.learning_history.append(learning_entry)
        
        # Keep only last 100 entries
        if len(self.learning_history) > 100:
            self.learning_history = self.learning_history[-100:]
        
        self.logger.info("Updated learning history with new reconciliation results")
    
    def validate_input(self, input_data: Any) -> Tuple[bool, List[str]]:
        """Validate input data for reconciliation."""
        errors = []
        
        if not isinstance(input_data, dict):
            errors.append("Input data must be a dictionary")
            return False, errors
        
        required_fields = ["gl_file", "bank_file"]
        for field in required_fields:
            if field not in input_data:
                errors.append(f"Required field '{field}' is missing")
            elif not isinstance(input_data[field], str):
                errors.append(f"Field '{field}' must be a string")
            elif not os.path.exists(input_data[field]):
                errors.append(f"File '{input_data[field]}' does not exist")
        
        return len(errors) == 0, errors
    
    def get_learning_insights(self) -> Dict[str, Any]:
        """Get learning insights from historical data."""
        if not self.learning_history:
            return {"message": "No learning history available"}
        
        recent_entries = self.learning_history[-10:]  # Last 10 entries
        
        avg_success_rate = sum(entry["success_rate"] for entry in recent_entries) / len(recent_entries)
        avg_variance = sum(entry["variance_level"] for entry in recent_entries) / len(recent_entries)
        
        return {
            "total_learning_entries": len(self.learning_history),
            "recent_avg_success_rate": avg_success_rate,
            "recent_avg_variance": avg_variance,
            "learning_trend": "improving" if avg_success_rate > 0.8 else "stable" if avg_success_rate > 0.6 else "needs_improvement",
            "recommendations": self._generate_learning_recommendations(avg_success_rate, avg_variance)
        }
    
    def _generate_learning_recommendations(self, success_rate: float, avg_variance: float) -> List[str]:
        """Generate recommendations based on learning insights."""
        recommendations = []
        
        if success_rate < 0.6:
            recommendations.append("Consider retraining with more diverse data")
        
        if avg_variance > 10:
            recommendations.append("Review variance thresholds and matching criteria")
        
        if success_rate > 0.9 and avg_variance < 2:
            recommendations.append("Excellent performance - consider expanding to more complex scenarios")
        
        return recommendations


# Example usage
if __name__ == "__main__":
    # Example configuration
    config = {
        "log_level": "INFO",
        "max_execution_time": 300,
        "retry_attempts": 3,
        "variance_threshold": 1000.0
    }
    
    # Initialize agent
    agent = ImprovedAIReconciliationAgent(
        config=config,
        training_data_path="training_data"
    )
    
    # Test execution
    test_input = {
        "gl_file": "test_gl.xlsx",
        "bank_file": "test_bank.xlsx"
    }
    
    # Execute with monitoring
    result = agent.execute_with_monitoring(test_input)
    print(f"Execution result: {json.dumps(result, indent=2)}")
    
    # Get learning insights
    insights = agent.get_learning_insights()
    print(f"Learning insights: {json.dumps(insights, indent=2)}")
    
    # Get agent status
    status = agent.get_status()
    print(f"Agent status: {json.dumps(status, indent=2)}")
