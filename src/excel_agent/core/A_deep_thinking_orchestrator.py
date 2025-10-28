#!/usr/bin/env python3
"""
Deep Thinking Orchestrator
Performs 10 rounds of deep thinking analysis on training documents and data review
using OpenAI API for advanced reasoning and pattern recognition
"""

import os
import json
import openai
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import pandas as pd
import openpyxl
import xlrd
from strands_base_agent import StrandsBaseAgent
from training_data_manager import TrainingDataManager

class DeepThinkingOrchestrator(StrandsBaseAgent):
    """
    Advanced orchestrator that performs deep thinking analysis using OpenAI API
    for comprehensive training document review and data analysis.
    """
    
    def __init__(self, 
                 config: Optional[Dict[str, Any]] = None,
                 training_data_path: Optional[str] = None,
                 openai_api_key: Optional[str] = None):
        """
        Initialize the deep thinking orchestrator.
        
        Args:
            config: Configuration dictionary
            training_data_path: Path to training data directory
            openai_api_key: OpenAI API key for deep thinking capabilities
        """
        super().__init__(
            name="DeepThinkingOrchestrator",
            config=config,
            training_data_path=training_data_path
        )
        
        # Initialize OpenAI
        self.openai_api_key = openai_api_key or os.getenv('OPENAI_API_KEY')
        if not self.openai_api_key:
            raise ValueError("OpenAI API key is required for deep thinking capabilities")
        
        openai.api_key = self.openai_api_key
        
        # Initialize training data manager
        self.tdm = TrainingDataManager(training_data_path)
        
        # Deep thinking configuration
        self.thinking_rounds = 10
        self.thinking_results = []
        self.deep_insights = {}
        self.pattern_analysis = {}
        
        self.logger.info("Deep Thinking Orchestrator initialized with OpenAI API")
    
    def process(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Process deep thinking analysis on training documents and data.
        
        Args:
            input_data: Dictionary containing analysis parameters
            
        Returns:
            Comprehensive deep thinking analysis results
        """
        try:
            self.logger.info("Starting deep thinking analysis process")
            
            # Phase 1: Deep thinking on training documents (10 rounds)
            training_analysis = self._perform_deep_training_analysis()
            
            # Phase 2: Deep review of NCB banking statement data
            ncb_analysis = self._perform_deep_ncb_analysis()
            
            # Phase 3: Deep review of Flex activity data
            flex_analysis = self._perform_deep_flex_analysis()
            
            # Phase 4: Synthesize all insights
            synthesis = self._synthesize_deep_insights(training_analysis, ncb_analysis, flex_analysis)
            
            # Phase 5: Generate comprehensive recommendations
            recommendations = self._generate_deep_recommendations(synthesis)
            
            return {
                "deep_thinking_results": {
                    "training_analysis": training_analysis,
                    "ncb_analysis": ncb_analysis,
                    "flex_analysis": flex_analysis,
                    "synthesis": synthesis,
                    "recommendations": recommendations
                },
                "thinking_rounds_completed": self.thinking_rounds,
                "analysis_timestamp": datetime.now().isoformat(),
                "total_insights_generated": len(self.deep_insights)
            }
            
        except Exception as e:
            self.logger.error(f"Error in deep thinking analysis: {str(e)}")
            raise
    
    def _perform_deep_training_analysis(self) -> Dict[str, Any]:
        """Perform 10 rounds of deep thinking analysis on training documents."""
        self.logger.info("🧠 Starting 10 rounds of deep thinking on training documents...")
        
        training_analysis = {
            "rounds_completed": 0,
            "insights_per_round": [],
            "cumulative_insights": [],
            "pattern_evolution": [],
            "confidence_scores": [],
            "key_discoveries": []
        }
        
        # Get training data
        op_manual = self.tdm.get_training_data("op_manual")
        historical_patterns = self.tdm.get_training_data("historical_patterns")
        reconciliation_rules = self.tdm.get_training_data("reconciliation_rules")
        
        for round_num in range(1, self.thinking_rounds + 1):
            self.logger.info(f"🔄 Deep Thinking Round {round_num}/10")
            
            # Create thinking prompt for this round
            thinking_prompt = self._create_thinking_prompt(round_num, op_manual, historical_patterns, reconciliation_rules)
            
            # Perform deep thinking using OpenAI
            round_insights = self._perform_openai_thinking(thinking_prompt, round_num)
            
            # Process and store insights
            training_analysis["insights_per_round"].append(round_insights)
            training_analysis["rounds_completed"] = round_num
            
            # Update cumulative insights
            self._update_cumulative_insights(round_insights, training_analysis)
            
            # Calculate confidence score for this round
            confidence = self._calculate_confidence_score(round_insights)
            training_analysis["confidence_scores"].append(confidence)
            
            self.logger.info(f"✅ Round {round_num} completed - Confidence: {confidence:.2f}")
        
        # Generate final insights
        training_analysis["final_insights"] = self._generate_final_training_insights(training_analysis)
        
        self.logger.info("🎯 Deep training analysis completed - 10 rounds finished")
        return training_analysis
    
    def _create_thinking_prompt(self, round_num: int, op_manual: Dict, historical_patterns: Dict, reconciliation_rules: Dict) -> str:
        """Create a sophisticated thinking prompt for each round."""
        
        base_prompt = f"""
You are an expert financial reconciliation analyst with deep expertise in GL accounting, bank statement analysis, and reconciliation processes. You are performing round {round_num} of 10 deep thinking sessions to analyze training documents and extract critical insights for improving reconciliation accuracy.

CONTEXT:
- You have access to OP manual data with GL account mappings and rules
- Historical patterns show common reconciliation challenges and solutions
- Reconciliation rules define matching criteria and validation standards
- Your goal is to extract increasingly sophisticated insights with each round

TRAINING DATA AVAILABLE:
OP Manual GL Accounts: {json.dumps(list(op_manual.get('gl_accounts', {}).keys()), indent=2)}
Historical Patterns: {len(historical_patterns.get('common_discrepancies', []))} discrepancy types identified
Reconciliation Rules: {len(reconciliation_rules.get('matching_criteria', {}))} matching criteria defined

THINKING FOCUS FOR ROUND {round_num}:
"""
        
        if round_num == 1:
            focus = "Focus on understanding the basic structure and relationships between GL accounts and bank activities. Identify fundamental patterns and rules."
        elif round_num == 2:
            focus = "Analyze timing differences and their impact on reconciliation accuracy. Look for patterns in when transactions are posted vs when they appear in bank statements."
        elif round_num == 3:
            focus = "Examine variance thresholds and their effectiveness. Consider how different GL accounts have different tolerance levels and why."
        elif round_num == 4:
            focus = "Investigate matching criteria and their reliability. Analyze which matching approaches work best for different types of transactions."
        elif round_num == 5:
            focus = "Look for edge cases and exceptions. Identify scenarios where standard rules might not apply and how to handle them."
        elif round_num == 6:
            focus = "Analyze historical failure patterns. Understand what causes reconciliation failures and how to prevent them."
        elif round_num == 7:
            focus = "Examine success patterns and best practices. Identify what makes reconciliations successful and how to replicate this."
        elif round_num == 8:
            focus = "Consider data quality issues and their impact. Analyze how data quality affects reconciliation accuracy."
        elif round_num == 9:
            focus = "Synthesize insights from previous rounds. Look for connections and relationships between different aspects of reconciliation."
        else:  # round_num == 10
            focus = "Generate final comprehensive insights and recommendations. Create actionable strategies for improving reconciliation processes."
        
        prompt = f"{base_prompt}\n{focus}\n\nProvide your deep analysis in the following format:\n1. Key Insights Discovered\n2. Pattern Analysis\n3. Rule Effectiveness Assessment\n4. Recommendations for Improvement\n5. Confidence Level (1-10)\n6. Questions for Further Investigation"
        
        return prompt
    
    def _perform_openai_thinking(self, prompt: str, round_num: int) -> Dict[str, Any]:
        """Perform deep thinking using OpenAI API."""
        try:
            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": "You are an expert financial reconciliation analyst with deep expertise in GL accounting, bank statement analysis, and reconciliation processes. Provide detailed, analytical responses with specific insights and actionable recommendations."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=2000,
                temperature=0.7,
                top_p=0.9
            )
            
            thinking_result = {
                "round": round_num,
                "timestamp": datetime.now().isoformat(),
                "raw_response": response.choices[0].message.content,
                "parsed_insights": self._parse_thinking_response(response.choices[0].message.content),
                "model_used": "gpt-4",
                "tokens_used": response.usage.total_tokens if hasattr(response, 'usage') else 0
            }
            
            return thinking_result
            
        except Exception as e:
            self.logger.error(f"Error in OpenAI thinking round {round_num}: {str(e)}")
            return {
                "round": round_num,
                "timestamp": datetime.now().isoformat(),
                "error": str(e),
                "parsed_insights": {}
            }
    
    def _parse_thinking_response(self, response: str) -> Dict[str, Any]:
        """Parse the thinking response to extract structured insights."""
        try:
            # Simple parsing - in a real implementation, this would be more sophisticated
            lines = response.split('\n')
            insights = {
                "key_insights": [],
                "pattern_analysis": [],
                "rule_assessment": [],
                "recommendations": [],
                "confidence_level": 5,
                "questions": []
            }
            
            current_section = None
            for line in lines:
                line = line.strip()
                if line.startswith('1.') or 'Key Insights' in line:
                    current_section = 'key_insights'
                elif line.startswith('2.') or 'Pattern Analysis' in line:
                    current_section = 'pattern_analysis'
                elif line.startswith('3.') or 'Rule Effectiveness' in line:
                    current_section = 'rule_assessment'
                elif line.startswith('4.') or 'Recommendations' in line:
                    current_section = 'recommendations'
                elif line.startswith('5.') or 'Confidence Level' in line:
                    current_section = 'confidence'
                elif line.startswith('6.') or 'Questions' in line:
                    current_section = 'questions'
                elif line and current_section and not line.startswith(('1.', '2.', '3.', '4.', '5.', '6.')):
                    if current_section == 'confidence':
                        try:
                            insights['confidence_level'] = float(line.split(':')[-1].strip())
                        except:
                            pass
                    else:
                        insights[current_section].append(line)
            
            return insights
            
        except Exception as e:
            self.logger.error(f"Error parsing thinking response: {str(e)}")
            return {"error": str(e)}
    
    def _update_cumulative_insights(self, round_insights: Dict[str, Any], training_analysis: Dict[str, Any]):
        """Update cumulative insights from each round."""
        if "parsed_insights" in round_insights:
            insights = round_insights["parsed_insights"]
            
            # Add to cumulative insights
            for key, value in insights.items():
                if key != "confidence_level" and isinstance(value, list):
                    if key not in training_analysis["cumulative_insights"]:
                        training_analysis["cumulative_insights"][key] = []
                    training_analysis["cumulative_insights"][key].extend(value)
    
    def _calculate_confidence_score(self, round_insights: Dict[str, Any]) -> float:
        """Calculate confidence score for a round."""
        if "parsed_insights" in round_insights and "confidence_level" in round_insights["parsed_insights"]:
            return round_insights["parsed_insights"]["confidence_level"]
        return 5.0  # Default confidence
    
    def _generate_final_training_insights(self, training_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Generate final insights from all training analysis rounds."""
        final_insights = {
            "total_rounds": training_analysis["rounds_completed"],
            "average_confidence": sum(training_analysis["confidence_scores"]) / len(training_analysis["confidence_scores"]) if training_analysis["confidence_scores"] else 0,
            "key_discoveries": [],
            "pattern_insights": [],
            "rule_insights": [],
            "recommendations": [],
            "synthesis": ""
        }
        
        # Extract unique insights
        cumulative = training_analysis["cumulative_insights"]
        final_insights["key_discoveries"] = list(set(cumulative.get("key_insights", [])))
        final_insights["pattern_insights"] = list(set(cumulative.get("pattern_analysis", [])))
        final_insights["rule_insights"] = list(set(cumulative.get("rule_assessment", [])))
        final_insights["recommendations"] = list(set(cumulative.get("recommendations", [])))
        
        # Generate synthesis
        final_insights["synthesis"] = self._generate_synthesis_text(final_insights)
        
        return final_insights
    
    def _generate_synthesis_text(self, insights: Dict[str, Any]) -> str:
        """Generate synthesis text from insights."""
        synthesis = f"""
Based on {insights['total_rounds']} rounds of deep thinking analysis with an average confidence of {insights['average_confidence']:.1f}/10, the following key insights have been discovered:

KEY DISCOVERIES ({len(insights['key_discoveries'])}):
{chr(10).join(f"- {discovery}" for discovery in insights['key_discoveries'][:10])}

PATTERN INSIGHTS ({len(insights['pattern_insights'])}):
{chr(10).join(f"- {pattern}" for pattern in insights['pattern_insights'][:10])}

RULE INSIGHTS ({len(insights['rule_insights'])}):
{chr(10).join(f"- {rule}" for rule in insights['rule_insights'][:10])}

RECOMMENDATIONS ({len(insights['recommendations'])}):
{chr(10).join(f"- {rec}" for rec in insights['recommendations'][:10])}
"""
        return synthesis
    
    def _perform_deep_ncb_analysis(self) -> Dict[str, Any]:
        """Perform deep analysis of NCB banking statement data."""
        self.logger.info("🏦 Starting deep analysis of NCB banking statement data...")
        
        ncb_analysis = {
            "files_analyzed": [],
            "deep_insights": [],
            "pattern_analysis": {},
            "anomaly_detection": [],
            "recommendations": []
        }
        
        # Find NCB bank files
        data_path = Path("data")
        ncb_files = list(data_path.glob("*NCB Bank*.xls")) + list(data_path.glob("*NCB Bank*.xlsx"))
        
        for ncb_file in ncb_files:
            self.logger.info(f"📊 Analyzing NCB file: {ncb_file.name}")
            
            # Analyze file with deep thinking
            file_analysis = self._analyze_ncb_file_with_thinking(ncb_file)
            ncb_analysis["files_analyzed"].append(file_analysis)
        
        # Generate comprehensive insights
        ncb_analysis["comprehensive_insights"] = self._generate_ncb_insights(ncb_analysis)
        
        self.logger.info("✅ NCB analysis completed")
        return ncb_analysis
    
    def _analyze_ncb_file_with_thinking(self, ncb_file: Path) -> Dict[str, Any]:
        """Analyze NCB file using deep thinking approach."""
        file_analysis = {
            "file_name": ncb_file.name,
            "file_path": str(ncb_file),
            "analysis_timestamp": datetime.now().isoformat(),
            "thinking_insights": [],
            "data_quality": {},
            "transaction_patterns": {},
            "anomalies": []
        }
        
        try:
            # Load file data
            if ncb_file.suffix == '.xlsx':
                wb = openpyxl.load_workbook(ncb_file)
                sheet = wb.active
                data = self._extract_sheet_data(sheet)
                wb.close()
            else:
                wb = xlrd.open_workbook(ncb_file)
                sheet = wb.sheet_by_index(0)
                data = self._extract_xls_sheet_data(sheet)
            
            # Perform deep thinking analysis on the data
            thinking_prompt = self._create_ncb_thinking_prompt(data, ncb_file.name)
            thinking_result = self._perform_openai_thinking(thinking_prompt, 0)  # Special round for NCB
            
            file_analysis["thinking_insights"] = thinking_result.get("parsed_insights", {})
            file_analysis["data_quality"] = self._assess_data_quality(data)
            file_analysis["transaction_patterns"] = self._analyze_transaction_patterns(data)
            file_analysis["anomalies"] = self._detect_anomalies(data)
            
        except Exception as e:
            self.logger.error(f"Error analyzing NCB file {ncb_file.name}: {str(e)}")
            file_analysis["error"] = str(e)
        
        return file_analysis
    
    def _create_ncb_thinking_prompt(self, data: Dict[str, Any], file_name: str) -> str:
        """Create thinking prompt for NCB analysis."""
        return f"""
You are analyzing NCB banking statement data from file: {file_name}

DATA SUMMARY:
- Total rows: {data.get('total_rows', 0)}
- Date range: {data.get('date_range', 'Unknown')}
- Transaction count: {len(data.get('transactions', []))}
- Total amount: ${data.get('total_amount', 0):,.2f}

PERFORM DEEP ANALYSIS:
1. Identify patterns in transaction types and amounts
2. Look for anomalies or unusual activity
3. Analyze data quality and completeness
4. Identify potential reconciliation challenges
5. Suggest improvements for data processing
6. Assess confidence in data accuracy

Provide detailed insights in structured format.
"""
    
    def _extract_sheet_data(self, sheet) -> Dict[str, Any]:
        """Extract data from XLSX sheet."""
        data = {
            "rows": [],
            "transactions": [],
            "total_rows": 0,
            "total_amount": 0.0,
            "date_range": None
        }
        
        dates = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data["rows"].append(list(row))
                data["total_rows"] += 1
                
                # Extract transaction data
                transaction = self._extract_transaction_from_row(row)
                if transaction:
                    data["transactions"].append(transaction)
                    data["total_amount"] += abs(transaction.get("amount", 0))
                    
                    if transaction.get("date"):
                        dates.append(transaction["date"])
        
        if dates:
            data["date_range"] = f"{min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}"
        
        return data
    
    def _extract_xls_sheet_data(self, sheet) -> Dict[str, Any]:
        """Extract data from XLS sheet."""
        data = {
            "rows": [],
            "transactions": [],
            "total_rows": 0,
            "total_amount": 0.0,
            "date_range": None
        }
        
        dates = []
        for row_index in range(sheet.nrows):
            row = sheet.row_values(row_index)
            if any(cell for cell in row):
                data["rows"].append(row)
                data["total_rows"] += 1
                
                # Extract transaction data
                transaction = self._extract_transaction_from_row(row)
                if transaction:
                    data["transactions"].append(transaction)
                    data["total_amount"] += abs(transaction.get("amount", 0))
                    
                    if transaction.get("date"):
                        dates.append(transaction["date"])
        
        if dates:
            data["date_range"] = f"{min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}"
        
        return data
    
    def _extract_transaction_from_row(self, row: List[Any]) -> Optional[Dict[str, Any]]:
        """Extract transaction data from a row."""
        try:
            transaction = {
                "date": None,
                "description": "",
                "amount": 0,
                "reference": ""
            }
            
            for cell in row:
                if isinstance(cell, datetime):
                    transaction["date"] = cell
                elif isinstance(cell, str) and len(cell) > 2:
                    if not transaction["description"]:
                        transaction["description"] = cell
                    else:
                        transaction["reference"] = cell
                elif isinstance(cell, (int, float)) and cell != 0:
                    transaction["amount"] = cell
            
            return transaction if transaction["amount"] != 0 else None
            
        except Exception:
            return None
    
    def _assess_data_quality(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Assess data quality of NCB data."""
        quality = {
            "completeness_score": 0.0,
            "consistency_score": 0.0,
            "accuracy_score": 0.0,
            "issues": []
        }
        
        total_transactions = len(data.get("transactions", []))
        if total_transactions == 0:
            quality["issues"].append("No transactions found")
            return quality
        
        # Check completeness
        complete_transactions = sum(1 for t in data["transactions"] if t.get("date") and t.get("amount") and t.get("description"))
        quality["completeness_score"] = complete_transactions / total_transactions
        
        # Check consistency
        amount_consistency = sum(1 for t in data["transactions"] if isinstance(t.get("amount"), (int, float)))
        quality["consistency_score"] = amount_consistency / total_transactions
        
        # Check accuracy (basic validation)
        valid_amounts = sum(1 for t in data["transactions"] if isinstance(t.get("amount"), (int, float)) and abs(t.get("amount", 0)) > 0)
        quality["accuracy_score"] = valid_amounts / total_transactions
        
        return quality
    
    def _analyze_transaction_patterns(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze transaction patterns in NCB data."""
        patterns = {
            "amount_distribution": {},
            "daily_patterns": {},
            "description_patterns": {},
            "frequency_analysis": {}
        }
        
        transactions = data.get("transactions", [])
        if not transactions:
            return patterns
        
        # Amount distribution
        amounts = [abs(t.get("amount", 0)) for t in transactions]
        if amounts:
            patterns["amount_distribution"] = {
                "min": min(amounts),
                "max": max(amounts),
                "avg": sum(amounts) / len(amounts),
                "median": sorted(amounts)[len(amounts) // 2]
            }
        
        # Daily patterns
        daily_totals = {}
        for t in transactions:
            if t.get("date"):
                day = t["date"].strftime("%Y-%m-%d")
                daily_totals[day] = daily_totals.get(day, 0) + abs(t.get("amount", 0))
        
        patterns["daily_patterns"] = daily_totals
        
        # Description patterns
        descriptions = [t.get("description", "") for t in transactions if t.get("description")]
        pattern_counts = {}
        for desc in descriptions:
            words = desc.lower().split()
            for word in words:
                if len(word) > 3:
                    pattern_counts[word] = pattern_counts.get(word, 0) + 1
        
        patterns["description_patterns"] = dict(sorted(pattern_counts.items(), key=lambda x: x[1], reverse=True)[:10])
        
        return patterns
    
    def _detect_anomalies(self, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Detect anomalies in NCB data."""
        anomalies = []
        transactions = data.get("transactions", [])
        
        if not transactions:
            return anomalies
        
        amounts = [abs(t.get("amount", 0)) for t in transactions]
        if amounts:
            avg_amount = sum(amounts) / len(amounts)
            threshold = avg_amount * 3  # 3x average as threshold
            
            for i, t in enumerate(transactions):
                amount = abs(t.get("amount", 0))
                if amount > threshold:
                    anomalies.append({
                        "type": "high_amount",
                        "transaction_index": i,
                        "amount": amount,
                        "description": t.get("description", ""),
                        "severity": "high" if amount > avg_amount * 5 else "medium"
                    })
        
        return anomalies
    
    def _generate_ncb_insights(self, ncb_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Generate comprehensive insights from NCB analysis."""
        insights = {
            "total_files_analyzed": len(ncb_analysis["files_analyzed"]),
            "data_quality_summary": {},
            "pattern_summary": {},
            "anomaly_summary": {},
            "recommendations": []
        }
        
        # Aggregate data quality metrics
        quality_scores = []
        for file_analysis in ncb_analysis["files_analyzed"]:
            if "data_quality" in file_analysis:
                quality_scores.append(file_analysis["data_quality"])
        
        if quality_scores:
            insights["data_quality_summary"] = {
                "avg_completeness": sum(q.get("completeness_score", 0) for q in quality_scores) / len(quality_scores),
                "avg_consistency": sum(q.get("consistency_score", 0) for q in quality_scores) / len(quality_scores),
                "avg_accuracy": sum(q.get("accuracy_score", 0) for q in quality_scores) / len(quality_scores)
            }
        
        # Generate recommendations
        insights["recommendations"] = [
            "Implement automated data quality validation",
            "Set up anomaly detection alerts",
            "Standardize transaction description formats",
            "Create data completeness checks"
        ]
        
        return insights
    
    def _perform_deep_flex_analysis(self) -> Dict[str, Any]:
        """Perform deep analysis of Flex activity data."""
        self.logger.info("📊 Starting deep analysis of Flex activity data...")
        
        flex_analysis = {
            "files_analyzed": [],
            "gl_accounts_analyzed": {},
            "deep_insights": [],
            "reconciliation_readiness": {},
            "recommendations": []
        }
        
        # Find Flex activity files
        data_path = Path("data")
        flex_files = list(data_path.glob("*Flex GL Activity*.xlsx")) + list(data_path.glob("*Reconciliation*.xlsx"))
        
        for flex_file in flex_files:
            self.logger.info(f"📊 Analyzing Flex file: {flex_file.name}")
            
            # Analyze file with deep thinking
            file_analysis = self._analyze_flex_file_with_thinking(flex_file)
            flex_analysis["files_analyzed"].append(file_analysis)
        
        # Generate comprehensive insights
        flex_analysis["comprehensive_insights"] = self._generate_flex_insights(flex_analysis)
        
        self.logger.info("✅ Flex analysis completed")
        return flex_analysis
    
    def _analyze_flex_file_with_thinking(self, flex_file: Path) -> Dict[str, Any]:
        """Analyze Flex file using deep thinking approach."""
        file_analysis = {
            "file_name": flex_file.name,
            "file_path": str(flex_file),
            "analysis_timestamp": datetime.now().isoformat(),
            "thinking_insights": [],
            "gl_accounts": {},
            "reconciliation_sheet": None,
            "data_quality": {},
            "anomalies": []
        }
        
        try:
            wb = openpyxl.load_workbook(flex_file)
            
            # Analyze each GL account sheet
            gl_accounts = ['74400', '74505', '74510', '74515', '74520', '74525', '74530', '74535', '74540', '74550', '74560', '74570']
            
            for gl_account in gl_accounts:
                if gl_account in wb.sheetnames:
                    sheet = wb[gl_account]
                    gl_analysis = self._analyze_gl_sheet_with_thinking(sheet, gl_account, flex_file.name)
                    file_analysis["gl_accounts"][gl_account] = gl_analysis
            
            # Check for reconciliation sheet
            reconciliation_sheets = [name for name in wb.sheetnames if "reconciliation" in name.lower() or "final" in name.lower()]
            if reconciliation_sheets:
                file_analysis["reconciliation_sheet"] = reconciliation_sheets[0]
            
            wb.close()
            
        except Exception as e:
            self.logger.error(f"Error analyzing Flex file {flex_file.name}: {str(e)}")
            file_analysis["error"] = str(e)
        
        return file_analysis
    
    def _analyze_gl_sheet_with_thinking(self, sheet, gl_account: str, file_name: str) -> Dict[str, Any]:
        """Analyze GL sheet using deep thinking approach."""
        gl_analysis = {
            "gl_account": gl_account,
            "file_name": file_name,
            "analysis_timestamp": datetime.now().isoformat(),
            "thinking_insights": [],
            "balance_data": {},
            "transaction_analysis": {},
            "data_quality": {},
            "anomalies": []
        }
        
        try:
            # Extract balance data
            balance_data = self._extract_gl_balance_data(sheet)
            gl_analysis["balance_data"] = balance_data
            
            # Perform deep thinking analysis
            thinking_prompt = self._create_flex_thinking_prompt(gl_account, balance_data, file_name)
            thinking_result = self._perform_openai_thinking(thinking_prompt, 0)  # Special round for Flex
            
            gl_analysis["thinking_insights"] = thinking_result.get("parsed_insights", {})
            
            # Analyze transactions
            gl_analysis["transaction_analysis"] = self._analyze_gl_transactions(sheet)
            
            # Assess data quality
            gl_analysis["data_quality"] = self._assess_gl_data_quality(balance_data)
            
            # Detect anomalies
            gl_analysis["anomalies"] = self._detect_gl_anomalies(balance_data)
            
        except Exception as e:
            self.logger.error(f"Error analyzing GL sheet {gl_account}: {str(e)}")
            gl_analysis["error"] = str(e)
        
        return gl_analysis
    
    def _create_flex_thinking_prompt(self, gl_account: str, balance_data: Dict[str, Any], file_name: str) -> str:
        """Create thinking prompt for Flex analysis."""
        return f"""
You are analyzing Flex GL Activity data for account {gl_account} from file: {file_name}

BALANCE DATA:
- Debits: ${balance_data.get('debits', 0):,.2f}
- Credits: ${balance_data.get('credits', 0):,.2f}
- Net Balance: ${balance_data.get('net_balance', 0):,.2f}
- Transaction Count: {balance_data.get('transaction_count', 0)}

PERFORM DEEP ANALYSIS:
1. Assess the balance accuracy and reasonableness
2. Identify potential reconciliation challenges
3. Look for data quality issues or anomalies
4. Analyze transaction patterns and timing
5. Suggest improvements for data processing
6. Assess readiness for reconciliation

Provide detailed insights in structured format.
"""
    
    def _extract_gl_balance_data(self, sheet) -> Dict[str, Any]:
        """Extract balance data from GL sheet."""
        balance_data = {
            "debits": 0.0,
            "credits": 0.0,
            "net_balance": 0.0,
            "transaction_count": 0
        }
        
        debits = []
        credits = []
        
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)):
                    if cell > 0:
                        debits.append(cell)
                    elif cell < 0:
                        credits.append(abs(cell))
        
        balance_data["debits"] = sum(debits)
        balance_data["credits"] = sum(credits)
        balance_data["net_balance"] = balance_data["debits"] - balance_data["credits"]
        balance_data["transaction_count"] = len(debits) + len(credits)
        
        return balance_data
    
    def _analyze_gl_transactions(self, sheet) -> Dict[str, Any]:
        """Analyze GL transactions."""
        analysis = {
            "transaction_count": 0,
            "amount_distribution": {},
            "daily_patterns": {},
            "description_patterns": {}
        }
        
        transactions = []
        for row in sheet.iter_rows(values_only=True):
            transaction = self._extract_transaction_from_row(row)
            if transaction:
                transactions.append(transaction)
        
        analysis["transaction_count"] = len(transactions)
        
        if transactions:
            amounts = [abs(t.get("amount", 0)) for t in transactions]
            analysis["amount_distribution"] = {
                "min": min(amounts),
                "max": max(amounts),
                "avg": sum(amounts) / len(amounts)
            }
        
        return analysis
    
    def _assess_gl_data_quality(self, balance_data: Dict[str, Any]) -> Dict[str, Any]:
        """Assess GL data quality."""
        quality = {
            "completeness_score": 1.0 if balance_data.get("transaction_count", 0) > 0 else 0.0,
            "consistency_score": 1.0 if balance_data.get("net_balance") == (balance_data.get("debits", 0) - balance_data.get("credits", 0)) else 0.0,
            "reasonableness_score": 1.0 if abs(balance_data.get("net_balance", 0)) < 1000000 else 0.5,  # Reasonable threshold
            "issues": []
        }
        
        if balance_data.get("transaction_count", 0) == 0:
            quality["issues"].append("No transactions found")
        
        if abs(balance_data.get("net_balance", 0)) > 1000000:
            quality["issues"].append("Unusually high balance")
        
        return quality
    
    def _detect_gl_anomalies(self, balance_data: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Detect anomalies in GL data."""
        anomalies = []
        
        net_balance = balance_data.get("net_balance", 0)
        if abs(net_balance) > 100000:  # High balance threshold
            anomalies.append({
                "type": "high_balance",
                "net_balance": net_balance,
                "severity": "high" if abs(net_balance) > 500000 else "medium"
            })
        
        if balance_data.get("transaction_count", 0) == 0 and (balance_data.get("debits", 0) > 0 or balance_data.get("credits", 0) > 0):
            anomalies.append({
                "type": "data_inconsistency",
                "description": "Non-zero balance with zero transactions",
                "severity": "high"
            })
        
        return anomalies
    
    def _generate_flex_insights(self, flex_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Generate comprehensive insights from Flex analysis."""
        insights = {
            "total_files_analyzed": len(flex_analysis["files_analyzed"]),
            "total_gl_accounts": 0,
            "reconciliation_readiness": {},
            "data_quality_summary": {},
            "recommendations": []
        }
        
        total_gl_accounts = 0
        quality_scores = []
        
        for file_analysis in flex_analysis["files_analyzed"]:
            total_gl_accounts += len(file_analysis.get("gl_accounts", {}))
            
            for gl_account, gl_data in file_analysis.get("gl_accounts", {}).items():
                if "data_quality" in gl_data:
                    quality_scores.append(gl_data["data_quality"])
        
        insights["total_gl_accounts"] = total_gl_accounts
        
        if quality_scores:
            insights["data_quality_summary"] = {
                "avg_completeness": sum(q.get("completeness_score", 0) for q in quality_scores) / len(quality_scores),
                "avg_consistency": sum(q.get("consistency_score", 0) for q in quality_scores) / len(quality_scores),
                "avg_reasonableness": sum(q.get("reasonableness_score", 0) for q in quality_scores) / len(quality_scores)
            }
        
        # Generate recommendations
        insights["recommendations"] = [
            "Implement GL balance validation rules",
            "Set up automated anomaly detection",
            "Create reconciliation readiness checks",
            "Standardize GL data formats"
        ]
        
        return insights
    
    def _synthesize_deep_insights(self, training_analysis: Dict[str, Any], ncb_analysis: Dict[str, Any], flex_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Synthesize insights from all analysis phases."""
        synthesis = {
            "synthesis_timestamp": datetime.now().isoformat(),
            "key_findings": [],
            "pattern_correlations": {},
            "reconciliation_insights": {},
            "data_quality_assessment": {},
            "strategic_recommendations": []
        }
        
        # Extract key findings
        synthesis["key_findings"] = [
            f"Training analysis completed {training_analysis['rounds_completed']} rounds with avg confidence {training_analysis['final_insights']['average_confidence']:.1f}/10",
            f"NCB analysis processed {ncb_analysis['comprehensive_insights']['total_files_analyzed']} files",
            f"Flex analysis processed {flex_analysis['comprehensive_insights']['total_files_analyzed']} files with {flex_analysis['comprehensive_insights']['total_gl_accounts']} GL accounts"
        ]
        
        # Generate strategic recommendations
        synthesis["strategic_recommendations"] = [
            "Implement deep thinking insights into reconciliation processes",
            "Use pattern recognition for improved matching accuracy",
            "Apply data quality improvements based on analysis findings",
            "Establish continuous learning feedback loops",
            "Create automated anomaly detection systems"
        ]
        
        return synthesis
    
    def _generate_deep_recommendations(self, synthesis: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Generate comprehensive recommendations based on deep analysis."""
        recommendations = [
            {
                "category": "Process Improvement",
                "priority": "High",
                "recommendation": "Implement 10-round deep thinking analysis for all reconciliation processes",
                "rationale": "Deep thinking analysis provides comprehensive insights and improves accuracy"
            },
            {
                "category": "Data Quality",
                "priority": "High",
                "recommendation": "Establish automated data quality validation based on deep analysis findings",
                "rationale": "Improved data quality reduces reconciliation errors and improves efficiency"
            },
            {
                "category": "Pattern Recognition",
                "priority": "Medium",
                "recommendation": "Implement AI-powered pattern recognition for transaction matching",
                "rationale": "Pattern recognition improves matching accuracy and reduces manual effort"
            },
            {
                "category": "Continuous Learning",
                "priority": "Medium",
                "recommendation": "Establish feedback loops for continuous improvement of reconciliation processes",
                "rationale": "Continuous learning ensures processes improve over time"
            },
            {
                "category": "Monitoring",
                "priority": "High",
                "recommendation": "Implement real-time monitoring of reconciliation processes with deep insights",
                "rationale": "Real-time monitoring enables proactive issue detection and resolution"
            }
        ]
        
        return recommendations
    
    def get_deep_insights_summary(self) -> Dict[str, Any]:
        """Get summary of deep insights generated."""
        return {
            "total_thinking_rounds": self.thinking_rounds,
            "insights_generated": len(self.deep_insights),
            "pattern_analysis_completed": len(self.pattern_analysis),
            "last_analysis": datetime.now().isoformat()
        }


# Example usage
if __name__ == "__main__":
    # Initialize deep thinking orchestrator
    config = {
        "log_level": "INFO",
        "max_execution_time": 1800,  # 30 minutes for deep analysis
        "retry_attempts": 3
    }
    
    orchestrator = DeepThinkingOrchestrator(
        config=config,
        training_data_path="training_data"
    )
    
    # Execute deep thinking analysis
    input_data = {
        "analysis_type": "comprehensive",
        "include_ncb_analysis": True,
        "include_flex_analysis": True
    }
    
    result = orchestrator.execute_with_monitoring(input_data)
    print(f"Deep thinking analysis completed: {json.dumps(result, indent=2)}")
    
    # Get insights summary
    summary = orchestrator.get_deep_insights_summary()
    print(f"Insights summary: {json.dumps(summary, indent=2)}")
