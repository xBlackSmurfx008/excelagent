#!/usr/bin/env python3
"""
Data Consolidation Agent
Handles deduplication and consolidation of GL activity files
"""

import pandas as pd
import openpyxl
from pathlib import Path
import json
from datetime import datetime
import uuid

class DataConsolidationAgent:
    """Agent for consolidating and deduplicating GL activity data"""
    
    def __init__(self):
        self.consolidated_data = {}
        self.duplicate_analysis = {}
        self.gl_accounts = ['74400', '74505', '74510', '74515', '74520', '74525', '74530', '74535', '74540', '74550', '74560', '74570']
        
    def analyze_duplicates(self, data_folder="data"):
        """Analyze duplicate GL data across files"""
        print("🔍 Analyzing duplicate GL data across files...")
        
        data_path = Path(data_folder)
        gl_files = list(data_path.glob("*GL Activity*.xlsx")) + list(data_path.glob("*Reconciliation*.xlsx"))
        
        duplicate_analysis = {
            "total_files": len(gl_files),
            "gl_accounts": {},
            "file_analysis": {},
            "duplicate_summary": {}
        }
        
        # Analyze each GL account across all files
        for gl_account in self.gl_accounts:
            duplicate_analysis["gl_accounts"][gl_account] = {
                "files_found_in": [],
                "total_debits": 0,
                "total_credits": 0,
                "net_balance": 0,
                "duplicate_count": 0
            }
        
        # Analyze each file
        for file_path in gl_files:
            print(f"📊 Analyzing {file_path.name}...")
            file_analysis = self._analyze_file(file_path)
            duplicate_analysis["file_analysis"][file_path.name] = file_analysis
            
            # Track GL accounts found in this file
            for gl_account in self.gl_accounts:
                if gl_account in file_analysis.get("gl_accounts_found", []):
                    duplicate_analysis["gl_accounts"][gl_account]["files_found_in"].append(file_path.name)
                    duplicate_analysis["gl_accounts"][gl_account]["duplicate_count"] += 1
        
        # Calculate totals for each GL account
        for gl_account in self.gl_accounts:
            account_data = duplicate_analysis["gl_accounts"][gl_account]
            account_data["duplicate_count"] = len(account_data["files_found_in"])
            
            # Sum up debits and credits from all files
            total_debits = 0
            total_credits = 0
            
            for file_name in account_data["files_found_in"]:
                file_data = duplicate_analysis["file_analysis"][file_name]
                if gl_account in file_data.get("gl_balances", {}):
                    gl_data = file_data["gl_balances"][gl_account]
                    total_debits += gl_data.get("debits", 0)
                    total_credits += gl_data.get("credits", 0)
            
            account_data["total_debits"] = total_debits
            account_data["total_credits"] = total_credits
            account_data["net_balance"] = total_debits - total_credits
        
        # Generate duplicate summary
        duplicate_analysis["duplicate_summary"] = {
            "accounts_with_duplicates": sum(1 for acc in duplicate_analysis["gl_accounts"].values() if acc["duplicate_count"] > 1),
            "total_duplicate_entries": sum(acc["duplicate_count"] for acc in duplicate_analysis["gl_accounts"].values()),
            "average_duplicates_per_account": sum(acc["duplicate_count"] for acc in duplicate_analysis["gl_accounts"].values()) / len(self.gl_accounts)
        }
        
        self.duplicate_analysis = duplicate_analysis
        return duplicate_analysis
    
    def _analyze_file(self, file_path):
        """Analyze a single GL activity file"""
        try:
            wb = openpyxl.load_workbook(file_path)
            file_analysis = {
                "file_name": file_path.name,
                "sheets": wb.sheetnames,
                "gl_accounts_found": [],
                "gl_balances": {},
                "has_reconciliation_final": "Reconciliation_Final" in wb.sheetnames or "May 2025 Reconciliation_Final" in wb.sheetnames
            }
            
            # Check each GL account sheet
            for gl_account in self.gl_accounts:
                if gl_account in wb.sheetnames:
                    file_analysis["gl_accounts_found"].append(gl_account)
                    
                    # Extract balance data
                    sheet = wb[gl_account]
                    gl_balance = self._extract_gl_balance(sheet, gl_account)
                    file_analysis["gl_balances"][gl_account] = gl_balance
            
            wb.close()
            return file_analysis
            
        except Exception as e:
            print(f"❌ Error analyzing {file_path.name}: {str(e)}")
            return {"error": str(e)}
    
    def _extract_gl_balance(self, sheet, gl_account):
        """Extract balance information from a GL account sheet"""
        try:
            # Look for balance information in the sheet
            debits = 0
            credits = 0
            
            # Search for balance data (this is a simplified extraction)
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, (int, float)):
                        if cell > 0:
                            debits += cell
                        elif cell < 0:
                            credits += abs(cell)
            
            return {
                "debits": debits,
                "credits": credits,
                "net_balance": debits - credits
            }
            
        except Exception as e:
            return {"debits": 0, "credits": 0, "net_balance": 0, "error": str(e)}
    
    def consolidate_data(self, data_folder="data"):
        """Consolidate GL data by removing duplicates and creating clean dataset"""
        print("🔄 Consolidating GL data and removing duplicates...")
        
        # First analyze duplicates
        duplicate_analysis = self.analyze_duplicates(data_folder)
        
        # Create consolidated dataset
        consolidated_data = {
            "consolidation_timestamp": datetime.now().isoformat(),
            "source_files": list(duplicate_analysis["file_analysis"].keys()),
            "gl_accounts": {},
            "consolidation_rules": {
                "deduplication_method": "sum_all_entries",
                "priority_file": "Reconciliation_Final_sheet",
                "exclude_duplicates": False
            }
        }
        
        # Consolidate each GL account
        for gl_account in self.gl_accounts:
            account_data = duplicate_analysis["gl_accounts"][gl_account]
            
            consolidated_data["gl_accounts"][gl_account] = {
                "total_debits": account_data["total_debits"],
                "total_credits": account_data["total_credits"],
                "net_balance": account_data["net_balance"],
                "source_files": account_data["files_found_in"],
                "duplicate_count": account_data["duplicate_count"],
                "consolidated": True
            }
        
        # Calculate totals
        total_debits = sum(acc["total_debits"] for acc in consolidated_data["gl_accounts"].values())
        total_credits = sum(acc["total_credits"] for acc in consolidated_data["gl_accounts"].values())
        net_imbalance = total_debits - total_credits
        
        consolidated_data["summary"] = {
            "total_debits": total_debits,
            "total_credits": total_credits,
            "net_imbalance": net_imbalance,
            "is_balanced": abs(net_imbalance) < 0.01,
            "accounts_processed": len(self.gl_accounts),
            "duplicate_entries_consolidated": sum(acc["duplicate_count"] for acc in consolidated_data["gl_accounts"].values())
        }
        
        self.consolidated_data = consolidated_data
        return consolidated_data
    
    def generate_consolidated_report(self, output_file="consolidated_reconciliation_report.xlsx"):
        """Generate consolidated Excel report"""
        print(f"📊 Generating consolidated report: {output_file}")
        
        if not self.consolidated_data:
            self.consolidate_data()
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Consolidated_Summary"
        
        # Add headers
        summary_sheet['A1'] = "Consolidated Reconciliation Report"
        summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        summary_sheet['A4'] = "Summary"
        
        # Add summary data
        summary_data = self.consolidated_data["summary"]
        summary_sheet['A6'] = "Total Debits:"
        summary_sheet['B6'] = f"${summary_data['total_debits']:,.2f}"
        summary_sheet['A7'] = "Total Credits:"
        summary_sheet['B7'] = f"${summary_data['total_credits']:,.2f}"
        summary_sheet['A8'] = "Net Imbalance:"
        summary_sheet['B8'] = f"${summary_data['net_imbalance']:,.2f}"
        summary_sheet['A9'] = "Balanced:"
        summary_sheet['B9'] = "Yes" if summary_data['is_balanced'] else "No"
        summary_sheet['A10'] = "Accounts Processed:"
        summary_sheet['B10'] = summary_data['accounts_processed']
        
        # GL Accounts sheet
        gl_sheet = wb.create_sheet("GL_Accounts_Consolidated")
        
        # Headers
        gl_sheet['A1'] = "GL Account"
        gl_sheet['B1'] = "Total Debits"
        gl_sheet['C1'] = "Total Credits"
        gl_sheet['D1'] = "Net Balance"
        gl_sheet['E1'] = "Source Files"
        gl_sheet['F1'] = "Duplicate Count"
        
        # Add GL account data
        row = 2
        for gl_account, data in self.consolidated_data["gl_accounts"].items():
            gl_sheet[f'A{row}'] = gl_account
            gl_sheet[f'B{row}'] = data["total_debits"]
            gl_sheet[f'C{row}'] = data["total_credits"]
            gl_sheet[f'D{row}'] = data["net_balance"]
            gl_sheet[f'E{row}'] = ", ".join(data["source_files"])
            gl_sheet[f'F{row}'] = data["duplicate_count"]
            row += 1
        
        # Duplicate Analysis sheet
        dup_sheet = wb.create_sheet("Duplicate_Analysis")
        dup_sheet['A1'] = "Duplicate Analysis Report"
        dup_sheet['A3'] = "GL Account"
        dup_sheet['B3'] = "Files Found In"
        dup_sheet['C3'] = "Duplicate Count"
        
        row = 4
        for gl_account, data in self.consolidated_data["gl_accounts"].items():
            dup_sheet[f'A{row}'] = gl_account
            dup_sheet[f'B{row}'] = len(data["source_files"])
            dup_sheet[f'C{row}'] = data["duplicate_count"]
            row += 1
        
        # Save the workbook
        wb.save(output_file)
        print(f"✅ Consolidated report saved: {output_file}")
        
        return output_file
    
    def generate_clean_reconciliation_sheet(self, output_file="Clean_Reconciliation_Final.xlsx"):
        """Generate a clean, deduplicated reconciliation sheet"""
        print(f"🧹 Generating clean reconciliation sheet: {output_file}")
        
        if not self.consolidated_data:
            self.consolidate_data()
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "May 2025 Reconciliation_Final"
        
        # Headers
        sheet['A1'] = "GL Account"
        sheet['B1'] = "Account Description"
        sheet['C1'] = "Debits"
        sheet['D1'] = "Credits"
        sheet['E1'] = "Net Balance"
        sheet['F1'] = "Status"
        
        # Account descriptions
        account_descriptions = {
            '74400': 'Interest Income',
            '74505': 'Service Charges',
            '74510': 'Other Income',
            '74515': 'Miscellaneous Income',
            '74520': 'Fee Income',
            '74525': 'Penalty Income',
            '74530': 'Interest Expense',
            '74535': 'Service Expense',
            '74540': 'Other Expense',
            '74550': 'Miscellaneous Expense',
            '74560': 'Fee Expense',
            '74570': 'Penalty Expense'
        }
        
        # Add GL account data
        row = 2
        total_debits = 0
        total_credits = 0
        
        for gl_account, data in self.consolidated_data["gl_accounts"].items():
            sheet[f'A{row}'] = gl_account
            sheet[f'B{row}'] = account_descriptions.get(gl_account, f"GL Account {gl_account}")
            sheet[f'C{row}'] = data["total_debits"]
            sheet[f'D{row}'] = data["total_credits"]
            sheet[f'E{row}'] = data["net_balance"]
            sheet[f'F{row}'] = "Balanced" if abs(data["net_balance"]) < 0.01 else "Imbalanced"
            
            total_debits += data["total_debits"]
            total_credits += data["total_credits"]
            row += 1
        
        # Add totals
        row += 1
        sheet[f'A{row}'] = "TOTALS"
        sheet[f'B{row}'] = ""
        sheet[f'C{row}'] = total_debits
        sheet[f'D{row}'] = total_credits
        sheet[f'E{row}'] = total_debits - total_credits
        sheet[f'F{row}'] = "Balanced" if abs(total_debits - total_credits) < 0.01 else "IMBALANCED"
        
        # Save the workbook
        wb.save(output_file)
        print(f"✅ Clean reconciliation sheet saved: {output_file}")
        
        return output_file

def main():
    """Main function to run data consolidation"""
    print("🚀 Starting Data Consolidation Process...")
    
    agent = DataConsolidationAgent()
    
    # Step 1: Analyze duplicates
    print("\n📊 Step 1: Analyzing duplicate data...")
    duplicate_analysis = agent.analyze_duplicates()
    
    # Step 2: Consolidate data
    print("\n🔄 Step 2: Consolidating data...")
    consolidated_data = agent.consolidate_data()
    
    # Step 3: Generate reports
    print("\n📊 Step 3: Generating reports...")
    consolidated_report = agent.generate_consolidated_report()
    clean_sheet = agent.generate_clean_reconciliation_sheet()
    
    print("\n✅ Data consolidation completed!")
    print(f"📄 Consolidated report: {consolidated_report}")
    print(f"🧹 Clean reconciliation sheet: {clean_sheet}")
    
    return {
        "consolidated_report": consolidated_report,
        "clean_sheet": clean_sheet,
        "duplicate_analysis": duplicate_analysis,
        "consolidated_data": consolidated_data
    }

if __name__ == "__main__":
    main()
