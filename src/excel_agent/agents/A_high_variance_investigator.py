#!/usr/bin/env python3
"""
High Variance Account Investigator
Investigates accounts with significant imbalances and variances
"""

import pandas as pd
import openpyxl
from pathlib import Path
import json
from datetime import datetime
import numpy as np

class HighVarianceInvestigator:
    """Agent for investigating high-variance GL accounts"""
    
    def __init__(self):
        self.investigation_results = {}
        self.high_variance_accounts = ['74505', '74520', '74530', '74560']
        self.variance_threshold = 1000000  # $1M threshold
        
    def investigate_high_variance_accounts(self, data_folder="data"):
        """Investigate accounts with high variances"""
        print("🔍 Investigating high-variance accounts...")
        
        data_path = Path(data_folder)
        gl_files = list(data_path.glob("*GL Activity*.xlsx")) + list(data_path.glob("*Reconciliation*.xlsx"))
        
        investigation_results = {
            "investigation_timestamp": datetime.now().isoformat(),
            "high_variance_accounts": {},
            "variance_analysis": {},
            "recommendations": {}
        }
        
        # Investigate each high-variance account
        for account in self.high_variance_accounts:
            print(f"📊 Investigating GL Account {account}...")
            account_analysis = self._investigate_account(account, gl_files)
            investigation_results["high_variance_accounts"][account] = account_analysis
        
        # Perform variance analysis
        investigation_results["variance_analysis"] = self._perform_variance_analysis(investigation_results["high_variance_accounts"])
        
        # Generate recommendations
        investigation_results["recommendations"] = self._generate_recommendations(investigation_results)
        
        self.investigation_results = investigation_results
        return investigation_results
    
    def _investigate_account(self, gl_account, gl_files):
        """Investigate a specific GL account across all files"""
        account_analysis = {
            "gl_account": gl_account,
            "files_containing_account": [],
            "balance_variations": [],
            "total_debits": 0,
            "total_credits": 0,
            "net_balance": 0,
            "variance_factors": [],
            "anomalies": []
        }
        
        # Analyze each file containing this account
        for file_path in gl_files:
            try:
                wb = openpyxl.load_workbook(file_path)
                if gl_account in wb.sheetnames:
                    account_analysis["files_containing_account"].append(file_path.name)
                    
                    # Extract detailed balance information
                    sheet = wb[gl_account]
                    balance_data = self._extract_detailed_balance(sheet, gl_account, file_path.name)
                    account_analysis["balance_variations"].append(balance_data)
                    
                    # Add to totals
                    account_analysis["total_debits"] += balance_data["debits"]
                    account_analysis["total_credits"] += balance_data["credits"]
                
                wb.close()
                
            except Exception as e:
                print(f"❌ Error analyzing {file_path.name} for account {gl_account}: {str(e)}")
        
        # Calculate net balance
        account_analysis["net_balance"] = account_analysis["total_debits"] - account_analysis["total_credits"]
        
        # Identify variance factors
        account_analysis["variance_factors"] = self._identify_variance_factors(account_analysis)
        
        # Detect anomalies
        account_analysis["anomalies"] = self._detect_anomalies(account_analysis)
        
        return account_analysis
    
    def _extract_detailed_balance(self, sheet, gl_account, file_name):
        """Extract detailed balance information from a GL account sheet"""
        balance_data = {
            "file_name": file_name,
            "debits": 0,
            "credits": 0,
            "net_balance": 0,
            "transaction_count": 0,
            "largest_debit": 0,
            "largest_credit": 0,
            "date_range": None
        }
        
        try:
            debits = []
            credits = []
            dates = []
            
            # Extract data from the sheet
            for row in sheet.iter_rows(values_only=True):
                for i, cell in enumerate(row):
                    if cell and isinstance(cell, (int, float)):
                        if cell > 0:
                            debits.append(cell)
                            balance_data["debits"] += cell
                        elif cell < 0:
                            credits.append(abs(cell))
                            balance_data["credits"] += abs(cell)
                    
                    # Look for date patterns
                    if isinstance(cell, datetime):
                        dates.append(cell)
            
            # Calculate statistics
            balance_data["net_balance"] = balance_data["debits"] - balance_data["credits"]
            balance_data["transaction_count"] = len(debits) + len(credits)
            balance_data["largest_debit"] = max(debits) if debits else 0
            balance_data["largest_credit"] = max(credits) if credits else 0
            
            if dates:
                balance_data["date_range"] = {
                    "earliest": min(dates).strftime('%Y-%m-%d'),
                    "latest": max(dates).strftime('%Y-%m-%d')
                }
            
        except Exception as e:
            print(f"⚠️ Error extracting balance from {file_name}: {str(e)}")
        
        return balance_data
    
    def _identify_variance_factors(self, account_analysis):
        """Identify factors contributing to variance"""
        factors = []
        
        # Check for multiple files with different balances
        if len(account_analysis["balance_variations"]) > 1:
            factors.append("Multiple files with different balances")
        
        # Check for large individual transactions
        for variation in account_analysis["balance_variations"]:
            if variation["largest_debit"] > self.variance_threshold:
                factors.append(f"Large debit transaction: ${variation['largest_debit']:,.2f} in {variation['file_name']}")
            if variation["largest_credit"] > self.variance_threshold:
                factors.append(f"Large credit transaction: ${variation['largest_credit']:,.2f} in {variation['file_name']}")
        
        # Check for date range inconsistencies
        date_ranges = [v["date_range"] for v in account_analysis["balance_variations"] if v["date_range"]]
        if len(set(str(dr) for dr in date_ranges)) > 1:
            factors.append("Inconsistent date ranges across files")
        
        return factors
    
    def _detect_anomalies(self, account_analysis):
        """Detect anomalies in account data"""
        anomalies = []
        
        # Check for zero balance in some files
        zero_balance_files = [v["file_name"] for v in account_analysis["balance_variations"] 
                            if v["net_balance"] == 0 and v["debits"] == 0 and v["credits"] == 0]
        if zero_balance_files:
            anomalies.append(f"Zero balance in files: {', '.join(zero_balance_files)}")
        
        # Check for extreme imbalances
        for variation in account_analysis["balance_variations"]:
            if abs(variation["net_balance"]) > self.variance_threshold * 2:
                anomalies.append(f"Extreme imbalance in {variation['file_name']}: ${variation['net_balance']:,.2f}")
        
        # Check for unusual transaction patterns
        for variation in account_analysis["balance_variations"]:
            if variation["transaction_count"] == 0 and (variation["debits"] > 0 or variation["credits"] > 0):
                anomalies.append(f"Non-zero balance with zero transactions in {variation['file_name']}")
        
        return anomalies
    
    def _perform_variance_analysis(self, high_variance_accounts):
        """Perform statistical variance analysis"""
        variance_analysis = {
            "total_variance": 0,
            "variance_by_account": {},
            "highest_variance_account": None,
            "variance_distribution": {}
        }
        
        max_variance = 0
        for account, data in high_variance_accounts.items():
            variance = abs(data["net_balance"])
            variance_analysis["variance_by_account"][account] = variance
            variance_analysis["total_variance"] += variance
            
            if variance > max_variance:
                max_variance = variance
                variance_analysis["highest_variance_account"] = account
        
        # Categorize variance levels
        variance_analysis["variance_distribution"] = {
            "extreme": sum(1 for v in variance_analysis["variance_by_account"].values() if v > self.variance_threshold * 2),
            "high": sum(1 for v in variance_analysis["variance_by_account"].values() if self.variance_threshold < v <= self.variance_threshold * 2),
            "moderate": sum(1 for v in variance_analysis["variance_by_account"].values() if v <= self.variance_threshold)
        }
        
        return variance_analysis
    
    def _generate_recommendations(self, investigation_results):
        """Generate recommendations based on investigation results"""
        recommendations = {
            "immediate_actions": [],
            "investigation_priorities": [],
            "process_improvements": [],
            "monitoring_recommendations": []
        }
        
        # Immediate actions
        for account, data in investigation_results["high_variance_accounts"].items():
            if data["net_balance"] > self.variance_threshold:
                recommendations["immediate_actions"].append(
                    f"Review GL Account {account}: ${data['net_balance']:,.2f} imbalance requires immediate attention"
                )
        
        # Investigation priorities
        variance_analysis = investigation_results["variance_analysis"]
        if variance_analysis["highest_variance_account"]:
            recommendations["investigation_priorities"].append(
                f"Priority 1: Investigate GL Account {variance_analysis['highest_variance_account']} "
                f"(${variance_analysis['variance_by_account'][variance_analysis['highest_variance_account']]:,.2f} variance)"
            )
        
        # Process improvements
        recommendations["process_improvements"].extend([
            "Implement automated duplicate detection in GL data processing",
            "Establish variance thresholds and alerting system",
            "Create standardized reconciliation templates",
            "Implement data validation rules for GL entries"
        ])
        
        # Monitoring recommendations
        recommendations["monitoring_recommendations"].extend([
            "Set up daily variance monitoring for high-risk accounts",
            "Implement automated reconciliation validation",
            "Create exception reporting for accounts exceeding variance thresholds",
            "Establish monthly variance review process"
        ])
        
        return recommendations
    
    def generate_investigation_report(self, output_file="high_variance_investigation_report.xlsx"):
        """Generate detailed investigation report"""
        print(f"📊 Generating investigation report: {output_file}")
        
        if not self.investigation_results:
            self.investigate_high_variance_accounts()
        
        wb = openpyxl.Workbook()
        
        # Executive Summary
        summary_sheet = wb.active
        summary_sheet.title = "Executive_Summary"
        
        summary_sheet['A1'] = "High Variance Account Investigation Report"
        summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Add summary data
        variance_analysis = self.investigation_results["variance_analysis"]
        summary_sheet['A4'] = "Total Variance:"
        summary_sheet['B4'] = f"${variance_analysis['total_variance']:,.2f}"
        summary_sheet['A5'] = "Highest Variance Account:"
        summary_sheet['B5'] = variance_analysis["highest_variance_account"]
        summary_sheet['A6'] = "Extreme Variance Accounts:"
        summary_sheet['B6'] = variance_analysis["variance_distribution"]["extreme"]
        summary_sheet['A7'] = "High Variance Accounts:"
        summary_sheet['B7'] = variance_analysis["variance_distribution"]["high"]
        
        # Account Details
        details_sheet = wb.create_sheet("Account_Details")
        
        # Headers
        headers = ["GL Account", "Total Debits", "Total Credits", "Net Balance", "Files", "Variance Factors", "Anomalies"]
        for i, header in enumerate(headers, 1):
            details_sheet.cell(row=1, column=i, value=header)
        
        # Add account data
        row = 2
        for account, data in self.investigation_results["high_variance_accounts"].items():
            details_sheet.cell(row=row, column=1, value=account)
            details_sheet.cell(row=row, column=2, value=data["total_debits"])
            details_sheet.cell(row=row, column=3, value=data["total_credits"])
            details_sheet.cell(row=row, column=4, value=data["net_balance"])
            details_sheet.cell(row=row, column=5, value=len(data["files_containing_account"]))
            details_sheet.cell(row=row, column=6, value="; ".join(data["variance_factors"]))
            details_sheet.cell(row=row, column=7, value="; ".join(data["anomalies"]))
            row += 1
        
        # Recommendations
        rec_sheet = wb.create_sheet("Recommendations")
        
        rec_sheet['A1'] = "Recommendations"
        row = 3
        
        for category, items in self.investigation_results["recommendations"].items():
            rec_sheet[f'A{row}'] = category.replace('_', ' ').title()
            row += 1
            for item in items:
                rec_sheet[f'A{row}'] = f"• {item}"
                row += 1
            row += 1
        
        # Save the workbook
        wb.save(output_file)
        print(f"✅ Investigation report saved: {output_file}")
        
        return output_file

def main():
    """Main function to run high variance investigation"""
    print("🔍 Starting High Variance Account Investigation...")
    
    investigator = HighVarianceInvestigator()
    
    # Run investigation
    investigation_results = investigator.investigate_high_variance_accounts()
    
    # Generate report
    report_file = investigator.generate_investigation_report()
    
    print("\n✅ High variance investigation completed!")
    print(f"📄 Investigation report: {report_file}")
    
    return {
        "investigation_results": investigation_results,
        "report_file": report_file
    }

if __name__ == "__main__":
    main()
