#!/usr/bin/env python3
"""
Discrepancy Discovery AI Agent
Focuses on finding WHERE discrepancies come from rather than forcing zero balance
This is a complex matching game of descriptions and numbers
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import os
from pathlib import Path

class DiscrepancyDiscoveryAgent:
    def __init__(self):
        self.name = "DiscrepancyDiscoveryAgent"
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
    def discover_discrepancies(self, gl_file_path, bank_file_path):
        """Main method to discover and explain discrepancies"""
        print("🔍 DISCREPANCY DISCOVERY AI AGENT")
        print("=" * 50)
        print("🎯 GOAL: Find WHERE discrepancies come from")
        print("🎮 COMPLEX MATCHING GAME: Descriptions & Numbers")
        print("=" * 50)
        
        # Load and analyze data
        gl_data = self._analyze_gl_activity(gl_file_path)
        bank_data = self._analyze_bank_statement(bank_file_path)
        
        # Perform discrepancy discovery
        discrepancies = self._discover_all_discrepancies(gl_data, bank_data)
        
        # Generate comprehensive report
        report = self._generate_discrepancy_report(discrepancies, gl_data, bank_data)
        
        return report
    
    def _analyze_gl_activity(self, file_path):
        """Analyze GL activity file"""
        print(f"📊 Analyzing GL Activity: {Path(file_path).name}")
        
        try:
            wb = openpyxl.load_workbook(file_path)
            gl_data = {}
            
            for sheet_name in wb.sheetnames:
                if 'GL' in sheet_name or 'Account' in sheet_name:
                    ws = wb[sheet_name]
                    gl_data[sheet_name] = self._parse_gl_sheet(ws)
            
            wb.close()
            return gl_data
            
        except Exception as e:
            print(f"❌ Error analyzing GL file: {str(e)}")
            return {}
    
    def _analyze_bank_statement(self, file_path):
        """Analyze bank statement file"""
        print(f"🏦 Analyzing Bank Statement: {Path(file_path).name}")
        
        try:
            import xlrd
            wb = xlrd.open_workbook(str(file_path))
            sheet = wb.sheet_by_index(0)
            
            transactions = []
            for row_idx in range(1, sheet.nrows):
                try:
                    account_number = str(sheet.cell_value(row_idx, 0))
                    post_date = sheet.cell_value(row_idx, 1)
                    check_number = sheet.cell_value(row_idx, 2)
                    description = str(sheet.cell_value(row_idx, 3)).strip()
                    debit = sheet.cell_value(row_idx, 4)
                    credit = sheet.cell_value(row_idx, 5)
                    
                    if post_date:
                        try:
                            date_tuple = xlrd.xldate_as_tuple(post_date, wb.datemode)
                            readable_date = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                        except:
                            readable_date = str(post_date)
                    else:
                        readable_date = ""
                    
                    debit_amount = float(debit) if debit else 0.0
                    credit_amount = float(credit) if credit else 0.0
                    
                    if debit_amount > 0 or credit_amount > 0:
                        transaction = {
                            "row": row_idx + 1,
                            "account_number": account_number,
                            "date": post_date,
                            "readable_date": readable_date,
                            "check_number": str(check_number) if check_number else "",
                            "description": description,
                            "debit": debit_amount,
                            "credit": credit_amount,
                            "type": "debit" if debit_amount > 0 else "credit",
                            "amount": debit_amount if debit_amount > 0 else credit_amount
                        }
                        transactions.append(transaction)
                
                except Exception as e:
                    print(f"⚠️ Error parsing row {row_idx + 1}: {str(e)}")
                    continue
            
            wb.release_resources()
            
            return {
                "file_name": Path(file_path).name,
                "account_number": transactions[0]["account_number"] if transactions else "Unknown",
                "transactions": transactions,
                "transaction_count": len(transactions)
            }
            
        except Exception as e:
            print(f"❌ Error analyzing bank statement: {str(e)}")
            return {"transactions": [], "transaction_count": 0}
    
    def _parse_gl_sheet(self, ws):
        """Parse a GL sheet"""
        transactions = []
        total_debits = 0.0
        total_credits = 0.0
        
        for row in range(2, ws.max_row + 1):
            try:
                description = ws.cell(row, 1).value
                debit = ws.cell(row, 2).value or 0
                credit = ws.cell(row, 3).value or 0
                
                if description and (debit > 0 or credit > 0):
                    transaction = {
                        "description": str(description),
                        "debit": float(debit),
                        "credit": float(credit),
                        "amount": float(debit) if debit > 0 else float(credit),
                        "type": "debit" if debit > 0 else "credit"
                    }
                    transactions.append(transaction)
                    total_debits += float(debit)
                    total_credits += float(credit)
            
            except Exception as e:
                continue
        
        return {
            "transactions": transactions,
            "total_debits": total_debits,
            "total_credits": total_credits,
            "balance": total_credits - total_debits,
            "transaction_count": len(transactions)
        }
    
    def _discover_all_discrepancies(self, gl_data, bank_data):
        """Discover all types of discrepancies"""
        print("🔍 DISCOVERING DISCREPANCIES...")
        print("-" * 40)
        
        discrepancies = {
            "perfect_matches": [],
            "close_matches": [],
            "unmatched_bank": [],
            "unmatched_gl": [],
            "timing_differences": [],
            "amount_differences": [],
            "description_mismatches": [],
            "date_differences": []
        }
        
        # Get all transactions
        bank_transactions = bank_data.get("transactions", [])
        all_gl_transactions = []
        
        for sheet_name, sheet_data in gl_data.items():
            for tx in sheet_data.get("transactions", []):
                tx["gl_sheet"] = sheet_name
                all_gl_transactions.append(tx)
        
        print(f"📊 Bank Transactions: {len(bank_transactions)}")
        print(f"📊 GL Transactions: {len(all_gl_transactions)}")
        
        # Perform matching analysis
        self._analyze_perfect_matches(bank_transactions, all_gl_transactions, discrepancies)
        self._analyze_close_matches(bank_transactions, all_gl_transactions, discrepancies)
        self._analyze_unmatched_items(bank_transactions, all_gl_transactions, discrepancies)
        self._analyze_timing_differences(bank_transactions, all_gl_transactions, discrepancies)
        self._analyze_amount_differences(bank_transactions, all_gl_transactions, discrepancies)
        self._analyze_description_mismatches(bank_transactions, all_gl_transactions, discrepancies)
        
        return discrepancies
    
    def _analyze_perfect_matches(self, bank_txs, gl_txs, discrepancies):
        """Find perfect matches"""
        print("🎯 Finding perfect matches...")
        
        for bank_tx in bank_txs:
            for gl_tx in gl_txs:
                if (abs(bank_tx["amount"] - gl_tx["amount"]) < 0.01 and
                    self._descriptions_match(bank_tx["description"], gl_tx["description"])):
                    
                    match = {
                        "bank_tx": bank_tx,
                        "gl_tx": gl_tx,
                        "confidence": "PERFECT",
                        "difference": 0.0,
                        "reason": "Exact amount and description match"
                    }
                    discrepancies["perfect_matches"].append(match)
                    break
        
        print(f"✅ Found {len(discrepancies['perfect_matches'])} perfect matches")
    
    def _analyze_close_matches(self, bank_txs, gl_txs, discrepancies):
        """Find close matches"""
        print("🔍 Finding close matches...")
        
        for bank_tx in bank_txs:
            for gl_tx in gl_txs:
                amount_diff = abs(bank_tx["amount"] - gl_tx["amount"])
                desc_similarity = self._description_similarity(bank_tx["description"], gl_tx["description"])
                
                if (amount_diff < 100.0 and desc_similarity > 0.3 and
                    not self._is_perfect_match(bank_tx, gl_tx, discrepancies)):
                    
                    match = {
                        "bank_tx": bank_tx,
                        "gl_tx": gl_tx,
                        "confidence": "CLOSE",
                        "difference": amount_diff,
                        "similarity": desc_similarity,
                        "reason": f"Close amount (${amount_diff:.2f} diff) and similar description"
                    }
                    discrepancies["close_matches"].append(match)
                    break
        
        print(f"✅ Found {len(discrepancies['close_matches'])} close matches")
    
    def _analyze_unmatched_items(self, bank_txs, gl_txs, discrepancies):
        """Find unmatched items"""
        print("🔍 Finding unmatched items...")
        
        matched_bank_indices = set()
        matched_gl_indices = set()
        
        # Mark matched items
        for match in discrepancies["perfect_matches"] + discrepancies["close_matches"]:
            for i, bank_tx in enumerate(bank_txs):
                if bank_tx == match["bank_tx"]:
                    matched_bank_indices.add(i)
            for i, gl_tx in enumerate(gl_txs):
                if gl_tx == match["gl_tx"]:
                    matched_gl_indices.add(i)
        
        # Find unmatched bank transactions
        for i, bank_tx in enumerate(bank_txs):
            if i not in matched_bank_indices:
                discrepancies["unmatched_bank"].append({
                    "transaction": bank_tx,
                    "reason": "No matching GL transaction found",
                    "suggestions": self._suggest_gl_matches(bank_tx, gl_txs)
                })
        
        # Find unmatched GL transactions
        for i, gl_tx in enumerate(gl_txs):
            if i not in matched_gl_indices:
                discrepancies["unmatched_gl"].append({
                    "transaction": gl_tx,
                    "reason": "No matching bank transaction found",
                    "suggestions": self._suggest_bank_matches(gl_tx, bank_txs)
                })
        
        print(f"⚠️ Found {len(discrepancies['unmatched_bank'])} unmatched bank transactions")
        print(f"⚠️ Found {len(discrepancies['unmatched_gl'])} unmatched GL transactions")
    
    def _analyze_timing_differences(self, bank_txs, gl_txs, discrepancies):
        """Analyze timing differences"""
        print("⏰ Analyzing timing differences...")
        
        # Look for transactions with same amount but different dates
        for bank_tx in bank_txs:
            for gl_tx in gl_txs:
                if (abs(bank_tx["amount"] - gl_tx["amount"]) < 0.01 and
                    bank_tx["readable_date"] != gl_tx.get("date", "")):
                    
                    timing_diff = {
                        "bank_tx": bank_tx,
                        "gl_tx": gl_tx,
                        "bank_date": bank_tx["readable_date"],
                        "gl_date": gl_tx.get("date", ""),
                        "amount": bank_tx["amount"],
                        "reason": "Same amount, different dates - likely timing difference"
                    }
                    discrepancies["timing_differences"].append(timing_diff)
        
        print(f"⏰ Found {len(discrepancies['timing_differences'])} timing differences")
    
    def _analyze_amount_differences(self, bank_txs, gl_txs, discrepancies):
        """Analyze amount differences"""
        print("💰 Analyzing amount differences...")
        
        for bank_tx in bank_txs:
            for gl_tx in gl_txs:
                amount_diff = abs(bank_tx["amount"] - gl_tx["amount"])
                if (amount_diff > 0.01 and amount_diff < 1000.0 and
                    self._descriptions_match(bank_tx["description"], gl_tx["description"])):
                    
                    amount_diff_item = {
                        "bank_tx": bank_tx,
                        "gl_tx": gl_tx,
                        "bank_amount": bank_tx["amount"],
                        "gl_amount": gl_tx["amount"],
                        "difference": amount_diff,
                        "reason": f"Same description, different amounts (${amount_diff:.2f} difference)"
                    }
                    discrepancies["amount_differences"].append(amount_diff_item)
        
        print(f"💰 Found {len(discrepancies['amount_differences'])} amount differences")
    
    def _analyze_description_mismatches(self, bank_txs, gl_txs, discrepancies):
        """Analyze description mismatches"""
        print("📝 Analyzing description mismatches...")
        
        for bank_tx in bank_txs:
            for gl_tx in gl_txs:
                if (abs(bank_tx["amount"] - gl_tx["amount"]) < 0.01 and
                    not self._descriptions_match(bank_tx["description"], gl_tx["description"])):
                    
                    desc_mismatch = {
                        "bank_tx": bank_tx,
                        "gl_tx": gl_tx,
                        "bank_description": bank_tx["description"],
                        "gl_description": gl_tx["description"],
                        "amount": bank_tx["amount"],
                        "reason": "Same amount, different descriptions"
                    }
                    discrepancies["description_mismatches"].append(desc_mismatch)
        
        print(f"📝 Found {len(discrepancies['description_mismatches'])} description mismatches")
    
    def _descriptions_match(self, desc1, desc2):
        """Check if descriptions match"""
        if not desc1 or not desc2:
            return False
        
        # Simple matching logic
        desc1_clean = desc1.upper().strip()
        desc2_clean = desc2.upper().strip()
        
        return desc1_clean == desc2_clean
    
    def _description_similarity(self, desc1, desc2):
        """Calculate description similarity"""
        if not desc1 or not desc2:
            return 0.0
        
        desc1_words = set(desc1.upper().split())
        desc2_words = set(desc2.upper().split())
        
        if not desc1_words or not desc2_words:
            return 0.0
        
        intersection = desc1_words.intersection(desc2_words)
        union = desc1_words.union(desc2_words)
        
        return len(intersection) / len(union) if union else 0.0
    
    def _is_perfect_match(self, bank_tx, gl_tx, discrepancies):
        """Check if this is already a perfect match"""
        for match in discrepancies["perfect_matches"]:
            if match["bank_tx"] == bank_tx and match["gl_tx"] == gl_tx:
                return True
        return False
    
    def _suggest_gl_matches(self, bank_tx, gl_txs):
        """Suggest possible GL matches for unmatched bank transaction"""
        suggestions = []
        for gl_tx in gl_txs:
            amount_diff = abs(bank_tx["amount"] - gl_tx["amount"])
            similarity = self._description_similarity(bank_tx["description"], gl_tx["description"])
            
            if amount_diff < 1000.0 or similarity > 0.2:
                suggestions.append({
                    "gl_transaction": gl_tx,
                    "amount_difference": amount_diff,
                    "similarity": similarity,
                    "reason": f"Amount diff: ${amount_diff:.2f}, Similarity: {similarity:.2f}"
                })
        
        return suggestions[:5]  # Top 5 suggestions
    
    def _suggest_bank_matches(self, gl_tx, bank_txs):
        """Suggest possible bank matches for unmatched GL transaction"""
        suggestions = []
        for bank_tx in bank_txs:
            amount_diff = abs(gl_tx["amount"] - bank_tx["amount"])
            similarity = self._description_similarity(gl_tx["description"], bank_tx["description"])
            
            if amount_diff < 1000.0 or similarity > 0.2:
                suggestions.append({
                    "bank_transaction": bank_tx,
                    "amount_difference": amount_diff,
                    "similarity": similarity,
                    "reason": f"Amount diff: ${amount_diff:.2f}, Similarity: {similarity:.2f}"
                })
        
        return suggestions[:5]  # Top 5 suggestions
    
    def _generate_discrepancy_report(self, discrepancies, gl_data, bank_data):
        """Generate comprehensive discrepancy report"""
        print("📊 GENERATING DISCREPANCY DISCOVERY REPORT")
        print("-" * 50)
        
        # Create Excel report
        excel_file = self._create_excel_discrepancy_report(discrepancies, gl_data, bank_data)
        
        # Create summary
        summary = {
            "timestamp": datetime.now().isoformat(),
            "perfect_matches": len(discrepancies["perfect_matches"]),
            "close_matches": len(discrepancies["close_matches"]),
            "unmatched_bank": len(discrepancies["unmatched_bank"]),
            "unmatched_gl": len(discrepancies["unmatched_gl"]),
            "timing_differences": len(discrepancies["timing_differences"]),
            "amount_differences": len(discrepancies["amount_differences"]),
            "description_mismatches": len(discrepancies["description_mismatches"]),
            "excel_file": excel_file
        }
        
        # Save JSON report
        with open(f"discrepancy_discovery_report_{self.timestamp}.json", 'w') as f:
            json.dump(summary, f, indent=2, default=str)
        
        print(f"✅ Discrepancy discovery report saved: {excel_file}")
        return summary
    
    def _create_excel_discrepancy_report(self, discrepancies, gl_data, bank_data):
        """Create Excel report with all discrepancies"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create sheets
        self._create_discrepancy_summary_sheet(wb, discrepancies)
        self._create_perfect_matches_sheet(wb, discrepancies)
        self._create_close_matches_sheet(wb, discrepancies)
        self._create_unmatched_bank_sheet(wb, discrepancies)
        self._create_unmatched_gl_sheet(wb, discrepancies)
        self._create_timing_differences_sheet(wb, discrepancies)
        self._create_amount_differences_sheet(wb, discrepancies)
        self._create_description_mismatches_sheet(wb, discrepancies)
        self._create_human_review_sheet(wb, discrepancies)
        
        # Save file
        excel_file = f"Discrepancy_Discovery_Report_{self.timestamp}.xlsx"
        wb.save(excel_file)
        
        return excel_file
    
    def _create_discrepancy_summary_sheet(self, wb, discrepancies):
        """Create summary sheet"""
        ws = wb.create_sheet("Discrepancy_Summary")
        
        # Title
        ws['A1'] = "🔍 DISCREPANCY DISCOVERY SUMMARY"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Summary data
        summary_data = [
            ["Perfect Matches", len(discrepancies["perfect_matches"]), "✅ EXACT"],
            ["Close Matches", len(discrepancies["close_matches"]), "⚠️ REVIEW"],
            ["Unmatched Bank", len(discrepancies["unmatched_bank"]), "❌ NEEDS ATTENTION"],
            ["Unmatched GL", len(discrepancies["unmatched_gl"]), "❌ NEEDS ATTENTION"],
            ["Timing Differences", len(discrepancies["timing_differences"]), "⏰ EXPECTED"],
            ["Amount Differences", len(discrepancies["amount_differences"]), "💰 REVIEW"],
            ["Description Mismatches", len(discrepancies["description_mismatches"]), "📝 REVIEW"]
        ]
        
        for i, (category, count, status) in enumerate(summary_data, 3):
            ws[f'A{i}'] = category
            ws[f'B{i}'] = count
            ws[f'C{i}'] = status
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
    
    def _create_perfect_matches_sheet(self, wb, discrepancies):
        """Create perfect matches sheet"""
        ws = wb.create_sheet("Perfect_Matches")
        
        # Headers
        headers = ["Bank Description", "Bank Amount", "GL Description", "GL Amount", "Confidence", "Reason"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
        
        # Data
        for row, match in enumerate(discrepancies["perfect_matches"], 2):
            ws.cell(row=row, column=1, value=match["bank_tx"]["description"])
            ws.cell(row=row, column=2, value=f"${match['bank_tx']['amount']:,.2f}")
            ws.cell(row=row, column=3, value=match["gl_tx"]["description"])
            ws.cell(row=row, column=4, value=f"${match['gl_tx']['amount']:,.2f}")
            ws.cell(row=row, column=5, value=match["confidence"])
            ws.cell(row=row, column=6, value=match["reason"])
        
        # Auto-fit columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_close_matches_sheet(self, wb, discrepancies):
        """Create close matches sheet"""
        ws = wb.create_sheet("Close_Matches")
        
        # Headers
        headers = ["Bank Description", "Bank Amount", "GL Description", "GL Amount", "Difference", "Similarity", "Reason"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        
        # Data
        for row, match in enumerate(discrepancies["close_matches"], 2):
            ws.cell(row=row, column=1, value=match["bank_tx"]["description"])
            ws.cell(row=row, column=2, value=f"${match['bank_tx']['amount']:,.2f}")
            ws.cell(row=row, column=3, value=match["gl_tx"]["description"])
            ws.cell(row=row, column=4, value=f"${match['gl_tx']['amount']:,.2f}")
            ws.cell(row=row, column=5, value=f"${match['difference']:,.2f}")
            ws.cell(row=row, column=6, value=f"{match['similarity']:.2f}")
            ws.cell(row=row, column=7, value=match["reason"])
        
        # Auto-fit columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_unmatched_bank_sheet(self, wb, discrepancies):
        """Create unmatched bank transactions sheet"""
        ws = wb.create_sheet("Unmatched_Bank")
        
        # Headers
        headers = ["Bank Description", "Amount", "Date", "Reason", "Top Suggestions"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Data
        for row, item in enumerate(discrepancies["unmatched_bank"], 2):
            ws.cell(row=row, column=1, value=item["transaction"]["description"])
            ws.cell(row=row, column=2, value=f"${item['transaction']['amount']:,.2f}")
            ws.cell(row=row, column=3, value=item["transaction"].get("readable_date", ""))
            ws.cell(row=row, column=4, value=item["reason"])
            
            # Top suggestions
            suggestions = item["suggestions"][:3]
            suggestion_text = "; ".join([f"GL: ${s['gl_transaction']['amount']:,.2f} (diff: ${s['amount_difference']:,.2f})" for s in suggestions])
            ws.cell(row=row, column=5, value=suggestion_text)
        
        # Auto-fit columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    def _create_unmatched_gl_sheet(self, wb, discrepancies):
        """Create unmatched GL transactions sheet"""
        ws = wb.create_sheet("Unmatched_GL")
        
        # Headers
        headers = ["GL Description", "Amount", "GL Sheet", "Reason", "Top Suggestions"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # Data
        for row, item in enumerate(discrepancies["unmatched_gl"], 2):
            ws.cell(row=row, column=1, value=item["transaction"]["description"])
            ws.cell(row=row, column=2, value=f"${item['transaction']['amount']:,.2f}")
            ws.cell(row=row, column=3, value=item["transaction"].get("gl_sheet", ""))
            ws.cell(row=row, column=4, value=item["reason"])
            
            # Top suggestions
            suggestions = item["suggestions"][:3]
            suggestion_text = "; ".join([f"Bank: ${s['bank_transaction']['amount']:,.2f} (diff: ${s['amount_difference']:,.2f})" for s in suggestions])
            ws.cell(row=row, column=5, value=suggestion_text)
        
        # Auto-fit columns
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    def _create_timing_differences_sheet(self, wb, discrepancies):
        """Create timing differences sheet"""
        ws = wb.create_sheet("Timing_Differences")
        
        # Headers
        headers = ["Bank Date", "Bank Description", "Amount", "GL Date", "GL Description", "Reason"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        
        # Data
        for row, item in enumerate(discrepancies["timing_differences"], 2):
            ws.cell(row=row, column=1, value=item["bank_date"])
            ws.cell(row=row, column=2, value=item["bank_tx"]["description"])
            ws.cell(row=row, column=3, value=f"${item['amount']:,.2f}")
            ws.cell(row=row, column=4, value=item["gl_date"])
            ws.cell(row=row, column=5, value=item["gl_tx"]["description"])
            ws.cell(row=row, column=6, value=item["reason"])
        
        # Auto-fit columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_amount_differences_sheet(self, wb, discrepancies):
        """Create amount differences sheet"""
        ws = wb.create_sheet("Amount_Differences")
        
        # Headers
        headers = ["Bank Description", "Bank Amount", "GL Description", "GL Amount", "Difference", "Reason"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
        
        # Data
        for row, item in enumerate(discrepancies["amount_differences"], 2):
            ws.cell(row=row, column=1, value=item["bank_tx"]["description"])
            ws.cell(row=row, column=2, value=f"${item['bank_amount']:,.2f}")
            ws.cell(row=row, column=3, value=item["gl_tx"]["description"])
            ws.cell(row=row, column=4, value=f"${item['gl_amount']:,.2f}")
            ws.cell(row=row, column=5, value=f"${item['difference']:,.2f}")
            ws.cell(row=row, column=6, value=item["reason"])
        
        # Auto-fit columns
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_description_mismatches_sheet(self, wb, discrepancies):
        """Create description mismatches sheet"""
        ws = wb.create_sheet("Description_Mismatches")
        
        # Headers
        headers = ["Bank Description", "GL Description", "Amount", "Reason"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
        
        # Data
        for row, item in enumerate(discrepancies["description_mismatches"], 2):
            ws.cell(row=row, column=1, value=item["bank_description"])
            ws.cell(row=row, column=2, value=item["gl_description"])
            ws.cell(row=row, column=3, value=f"${item['amount']:,.2f}")
            ws.cell(row=row, column=4, value=item["reason"])
        
        # Auto-fit columns
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    def _create_human_review_sheet(self, wb, discrepancies):
        """Create human review sheet"""
        ws = wb.create_sheet("Human_Review_Needed")
        
        # Title
        ws['A1'] = "👤 ITEMS REQUIRING HUMAN REVIEW"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Summary
        ws['A3'] = "Priority Items for Human Review:"
        ws['A3'].font = Font(bold=True, size=14)
        
        review_items = [
            f"❌ {len(discrepancies['unmatched_bank'])} Unmatched Bank Transactions - Need GL matches",
            f"❌ {len(discrepancies['unmatched_gl'])} Unmatched GL Transactions - Need bank matches",
            f"💰 {len(discrepancies['amount_differences'])} Amount Differences - Need verification",
            f"📝 {len(discrepancies['description_mismatches'])} Description Mismatches - Need review",
            f"⚠️ {len(discrepancies['close_matches'])} Close Matches - Need confirmation"
        ]
        
        for i, item in enumerate(review_items, 5):
            ws[f'A{i}'] = item
        
        ws['A11'] = "🎯 RECOMMENDATIONS:"
        ws['A11'].font = Font(bold=True, size=14)
        
        recommendations = [
            "1. Review unmatched bank transactions and find corresponding GL entries",
            "2. Review unmatched GL transactions and find corresponding bank entries",
            "3. Verify amount differences - may be fees, adjustments, or errors",
            "4. Check description mismatches - may be different naming conventions",
            "5. Confirm close matches are actually the same transaction",
            "6. Document any transactions that cannot be matched",
            "7. Create manual reconciliation entries for unmatched items"
        ]
        
        for i, rec in enumerate(recommendations, 13):
            ws[f'A{i}'] = rec
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 80

def main():
    """Run discrepancy discovery"""
    print("🔍 DISCREPANCY DISCOVERY AI AGENT")
    print("=" * 40)
    print("🎯 Focus: Find WHERE discrepancies come from")
    print("🎮 Complex matching game: Descriptions & Numbers")
    print("=" * 40)
    
    agent = DiscrepancyDiscoveryAgent()
    
    # File paths
    gl_file = 'Data/05 May 2025 Reconciliation and Flex GL Activity.xlsx'
    bank_file = 'Data/NCB Bank Activity 5-1 to 5-31 Support for May 2025 Rec.xls'
    
    # Run discrepancy discovery
    report = agent.discover_discrepancies(gl_file, bank_file)
    
    print(f"\n🎉 DISCREPANCY DISCOVERY COMPLETE!")
    print(f"📊 Excel Report: {report['excel_file']}")
    print(f"✅ Perfect Matches: {report['perfect_matches']}")
    print(f"⚠️ Close Matches: {report['close_matches']}")
    print(f"❌ Unmatched Bank: {report['unmatched_bank']}")
    print(f"❌ Unmatched GL: {report['unmatched_gl']}")
    print(f"⏰ Timing Differences: {report['timing_differences']}")
    print(f"💰 Amount Differences: {report['amount_differences']}")
    print(f"📝 Description Mismatches: {report['description_mismatches']}")

if __name__ == "__main__":
    main()
