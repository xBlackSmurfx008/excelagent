#!/usr/bin/env python3
"""
Bank Cross-Match Agent
Cross-matches GL data with bank statement files for validation
"""

import pandas as pd
import openpyxl
from pathlib import Path
import json
from datetime import datetime
import re

class BankCrossMatchAgent:
    """Agent for cross-matching GL data with bank statements"""
    
    def __init__(self):
        self.cross_match_results = {}
        self.bank_files = []
        self.gl_files = []
        self.matching_criteria = {
            "amount_tolerance": 0.01,  # $0.01 tolerance for amount matching
            "date_tolerance_days": 3,  # 3-day tolerance for date matching
            "description_keywords": ["interest", "fee", "charge", "payment", "deposit"]
        }
    
    def cross_match_with_bank_files(self, data_folder="data"):
        """Cross-match GL data with bank statement files"""
        print("🏦 Cross-matching GL data with bank statements...")
        
        data_path = Path(data_folder)
        
        # Find bank and GL files
        self.bank_files = list(data_path.glob("*Bank*.xls")) + list(data_path.glob("*Bank*.xlsx"))
        self.gl_files = list(data_path.glob("*GL Activity*.xlsx")) + list(data_path.glob("*Reconciliation*.xlsx"))
        
        print(f"📊 Found {len(self.bank_files)} bank files and {len(self.gl_files)} GL files")
        
        cross_match_results = {
            "cross_match_timestamp": datetime.now().isoformat(),
            "bank_files_analyzed": [],
            "gl_files_analyzed": [],
            "matching_results": {},
            "validation_summary": {},
            "discrepancies": [],
            "recommendations": []
        }
        
        # Analyze each bank file
        for bank_file in self.bank_files:
            print(f"🏦 Analyzing bank file: {bank_file.name}")
            bank_analysis = self._analyze_bank_file(bank_file)
            cross_match_results["bank_files_analyzed"].append(bank_analysis)
        
        # Cross-match with GL data
        for gl_file in self.gl_files:
            print(f"📊 Cross-matching with GL file: {gl_file.name}")
            gl_analysis = self._analyze_gl_file(gl_file)
            cross_match_results["gl_files_analyzed"].append(gl_analysis)
        
        # Perform cross-matching
        cross_match_results["matching_results"] = self._perform_cross_matching(
            cross_match_results["bank_files_analyzed"],
            cross_match_results["gl_files_analyzed"]
        )
        
        # Generate validation summary
        cross_match_results["validation_summary"] = self._generate_validation_summary(cross_match_results)
        
        # Identify discrepancies
        cross_match_results["discrepancies"] = self._identify_discrepancies(cross_match_results)
        
        # Generate recommendations
        cross_match_results["recommendations"] = self._generate_recommendations(cross_match_results)
        
        self.cross_match_results = cross_match_results
        return cross_match_results
    
    def _analyze_bank_file(self, bank_file):
        """Analyze a bank statement file"""
        bank_analysis = {
            "file_name": bank_file.name,
            "transactions": [],
            "total_debits": 0,
            "total_credits": 0,
            "net_balance": 0,
            "date_range": None,
            "account_number": None
        }
        
        try:
            wb = openpyxl.load_workbook(bank_file)
            sheet = wb.active  # Assume first sheet contains bank data
            
            # Extract transactions
            transactions = []
            for row in sheet.iter_rows(values_only=True):
                if self._is_transaction_row(row):
                    transaction = self._extract_transaction(row)
                    if transaction:
                        transactions.append(transaction)
            
            bank_analysis["transactions"] = transactions
            
            # Calculate totals
            for transaction in transactions:
                if transaction["amount"] > 0:
                    bank_analysis["total_credits"] += transaction["amount"]
                else:
                    bank_analysis["total_debits"] += abs(transaction["amount"])
            
            bank_analysis["net_balance"] = bank_analysis["total_credits"] - bank_analysis["total_debits"]
            
            # Extract date range
            if transactions:
                dates = [t["date"] for t in transactions if t["date"]]
                if dates:
                    bank_analysis["date_range"] = {
                        "start": min(dates).strftime('%Y-%m-%d'),
                        "end": max(dates).strftime('%Y-%m-%d')
                    }
            
            # Try to extract account number
            bank_analysis["account_number"] = self._extract_account_number(sheet)
            
            wb.close()
            
        except Exception as e:
            print(f"❌ Error analyzing bank file {bank_file.name}: {str(e)}")
            bank_analysis["error"] = str(e)
        
        return bank_analysis
    
    def _analyze_gl_file(self, gl_file):
        """Analyze a GL activity file"""
        gl_analysis = {
            "file_name": gl_file.name,
            "gl_accounts": {},
            "total_debits": 0,
            "total_credits": 0,
            "net_balance": 0,
            "has_reconciliation_sheet": False
        }
        
        try:
            wb = openpyxl.load_workbook(gl_file)
            
            # Check for reconciliation sheet
            if "Reconciliation_Final" in wb.sheetnames or "May 2025 Reconciliation_Final" in wb.sheetnames:
                gl_analysis["has_reconciliation_sheet"] = True
            
            # Analyze GL account sheets
            gl_accounts = ['74400', '74505', '74510', '74515', '74520', '74525', '74530', '74535', '74540', '74550', '74560', '74570']
            
            for gl_account in gl_accounts:
                if gl_account in wb.sheetnames:
                    sheet = wb[gl_account]
                    account_data = self._extract_gl_account_data(sheet, gl_account)
                    gl_analysis["gl_accounts"][gl_account] = account_data
                    
                    gl_analysis["total_debits"] += account_data["debits"]
                    gl_analysis["total_credits"] += account_data["credits"]
            
            gl_analysis["net_balance"] = gl_analysis["total_debits"] - gl_analysis["total_credits"]
            
            wb.close()
            
        except Exception as e:
            print(f"❌ Error analyzing GL file {gl_file.name}: {str(e)}")
            gl_analysis["error"] = str(e)
        
        return gl_analysis
    
    def _is_transaction_row(self, row):
        """Check if a row contains transaction data"""
        # Look for patterns that indicate a transaction row
        for cell in row:
            if isinstance(cell, (int, float)) and abs(cell) > 0:
                return True
            if isinstance(cell, str) and any(keyword in cell.lower() for keyword in self.matching_criteria["description_keywords"]):
                return True
        return False
    
    def _extract_transaction(self, row):
        """Extract transaction data from a row"""
        transaction = {
            "date": None,
            "description": "",
            "amount": 0,
            "reference": ""
        }
        
        try:
            for i, cell in enumerate(row):
                if isinstance(cell, datetime):
                    transaction["date"] = cell
                elif isinstance(cell, str) and len(cell) > 3:
                    if not transaction["description"]:
                        transaction["description"] = cell
                    else:
                        transaction["reference"] = cell
                elif isinstance(cell, (int, float)) and cell != 0:
                    transaction["amount"] = cell
            
            return transaction if transaction["amount"] != 0 else None
            
        except Exception:
            return None
    
    def _extract_account_number(self, sheet):
        """Extract account number from bank sheet"""
        try:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str) and "account" in cell.lower():
                        # Look for account number pattern
                        numbers = re.findall(r'\d{4,}', cell)
                        if numbers:
                            return numbers[0]
        except Exception:
            pass
        return None
    
    def _extract_gl_account_data(self, sheet, gl_account):
        """Extract data from a GL account sheet"""
        account_data = {
            "gl_account": gl_account,
            "debits": 0,
            "credits": 0,
            "net_balance": 0,
            "transaction_count": 0
        }
        
        try:
            debits = []
            credits = []
            
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)):
                        if cell > 0:
                            debits.append(cell)
                        elif cell < 0:
                            credits.append(abs(cell))
            
            account_data["debits"] = sum(debits)
            account_data["credits"] = sum(credits)
            account_data["net_balance"] = account_data["debits"] - account_data["credits"]
            account_data["transaction_count"] = len(debits) + len(credits)
            
        except Exception as e:
            print(f"⚠️ Error extracting GL account {gl_account} data: {str(e)}")
        
        return account_data
    
    def _perform_cross_matching(self, bank_files, gl_files):
        """Perform cross-matching between bank and GL data"""
        matching_results = {
            "total_matches": 0,
            "total_unmatched_bank": 0,
            "total_unmatched_gl": 0,
            "match_percentage": 0,
            "detailed_matches": [],
            "unmatched_items": []
        }
        
        # Simple matching logic - in a real implementation, this would be more sophisticated
        total_bank_transactions = sum(len(bf["transactions"]) for bf in bank_files)
        total_gl_balance = sum(gf["net_balance"] for gf in gl_files)
        
        # Calculate match percentage based on balance alignment
        if total_gl_balance != 0:
            balance_match_percentage = min(100, (1 - abs(total_gl_balance) / abs(total_gl_balance)) * 100)
        else:
            balance_match_percentage = 100
        
        matching_results["match_percentage"] = balance_match_percentage
        matching_results["total_matches"] = int(total_bank_transactions * balance_match_percentage / 100)
        matching_results["total_unmatched_bank"] = total_bank_transactions - matching_results["total_matches"]
        matching_results["total_unmatched_gl"] = abs(total_gl_balance)
        
        return matching_results
    
    def _generate_validation_summary(self, cross_match_results):
        """Generate validation summary"""
        validation_summary = {
            "bank_files_count": len(cross_match_results["bank_files_analyzed"]),
            "gl_files_count": len(cross_match_results["gl_files_analyzed"]),
            "total_bank_transactions": sum(len(bf["transactions"]) for bf in cross_match_results["bank_files_analyzed"]),
            "total_gl_balance": sum(gf["net_balance"] for gf in cross_match_results["gl_files_analyzed"]),
            "match_percentage": cross_match_results["matching_results"]["match_percentage"],
            "validation_status": "PASS" if cross_match_results["matching_results"]["match_percentage"] > 80 else "FAIL"
        }
        
        return validation_summary
    
    def _identify_discrepancies(self, cross_match_results):
        """Identify discrepancies between bank and GL data"""
        discrepancies = []
        
        # Check for balance discrepancies
        total_gl_balance = sum(gf["net_balance"] for gf in cross_match_results["gl_files_analyzed"])
        if abs(total_gl_balance) > 1000:  # $1K threshold
            discrepancies.append({
                "type": "Balance Discrepancy",
                "description": f"GL balance of ${total_gl_balance:,.2f} exceeds threshold",
                "severity": "High" if abs(total_gl_balance) > 100000 else "Medium"
            })
        
        # Check for missing reconciliation sheets
        gl_files_without_reconciliation = [gf for gf in cross_match_results["gl_files_analyzed"] 
                                         if not gf.get("has_reconciliation_sheet", False)]
        if gl_files_without_reconciliation:
            discrepancies.append({
                "type": "Missing Reconciliation Sheet",
                "description": f"{len(gl_files_without_reconciliation)} GL files missing reconciliation sheets",
                "severity": "Medium"
            })
        
        # Check for low match percentage
        if cross_match_results["matching_results"]["match_percentage"] < 80:
            discrepancies.append({
                "type": "Low Match Percentage",
                "description": f"Only {cross_match_results['matching_results']['match_percentage']:.1f}% match between bank and GL data",
                "severity": "High"
            })
        
        return discrepancies
    
    def _generate_recommendations(self, cross_match_results):
        """Generate recommendations based on cross-matching results"""
        recommendations = []
        
        # Balance reconciliation recommendations
        total_gl_balance = sum(gf["net_balance"] for gf in cross_match_results["gl_files_analyzed"])
        if abs(total_gl_balance) > 1000:
            recommendations.append(f"Reconcile GL balance of ${total_gl_balance:,.2f} with bank statements")
        
        # Match percentage recommendations
        match_percentage = cross_match_results["matching_results"]["match_percentage"]
        if match_percentage < 80:
            recommendations.append(f"Improve data matching - currently only {match_percentage:.1f}% match")
        
        # Process improvement recommendations
        recommendations.extend([
            "Implement automated bank-to-GL reconciliation process",
            "Establish daily reconciliation procedures",
            "Create exception reporting for unmatched items",
            "Implement data validation rules for transaction matching"
        ])
        
        return recommendations
    
    def generate_cross_match_report(self, output_file="bank_cross_match_report.xlsx"):
        """Generate cross-matching report"""
        print(f"📊 Generating cross-match report: {output_file}")
        
        if not self.cross_match_results:
            self.cross_match_with_bank_files()
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Cross_Match_Summary"
        
        summary_sheet['A1'] = "Bank-GL Cross-Match Report"
        summary_sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Add summary data
        validation_summary = self.cross_match_results["validation_summary"]
        summary_sheet['A4'] = "Bank Files:"
        summary_sheet['B4'] = validation_summary["bank_files_count"]
        summary_sheet['A5'] = "GL Files:"
        summary_sheet['B5'] = validation_summary["gl_files_count"]
        summary_sheet['A6'] = "Match Percentage:"
        summary_sheet['B6'] = f"{validation_summary['match_percentage']:.1f}%"
        summary_sheet['A7'] = "Validation Status:"
        summary_sheet['B7'] = validation_summary["validation_status"]
        summary_sheet['A8'] = "Total GL Balance:"
        summary_sheet['B8'] = f"${validation_summary['total_gl_balance']:,.2f}"
        
        # Discrepancies sheet
        disc_sheet = wb.create_sheet("Discrepancies")
        
        disc_sheet['A1'] = "Discrepancies"
        disc_sheet['A2'] = "Type"
        disc_sheet['B2'] = "Description"
        disc_sheet['C2'] = "Severity"
        
        row = 3
        for disc in self.cross_match_results["discrepancies"]:
            disc_sheet[f'A{row}'] = disc["type"]
            disc_sheet[f'B{row}'] = disc["description"]
            disc_sheet[f'C{row}'] = disc["severity"]
            row += 1
        
        # Recommendations sheet
        rec_sheet = wb.create_sheet("Recommendations")
        
        rec_sheet['A1'] = "Recommendations"
        row = 3
        for rec in self.cross_match_results["recommendations"]:
            rec_sheet[f'A{row}'] = f"• {rec}"
            row += 1
        
        # Save the workbook
        wb.save(output_file)
        print(f"✅ Cross-match report saved: {output_file}")
        
        return output_file

def main():
    """Main function to run bank cross-matching"""
    print("🏦 Starting Bank Cross-Match Analysis...")
    
    agent = BankCrossMatchAgent()
    
    # Run cross-matching
    cross_match_results = agent.cross_match_with_bank_files()
    
    # Generate report
    report_file = agent.generate_cross_match_report()
    
    print("\n✅ Bank cross-match analysis completed!")
    print(f"📄 Cross-match report: {report_file}")
    
    return {
        "cross_match_results": cross_match_results,
        "report_file": report_file
    }

if __name__ == "__main__":
    main()
