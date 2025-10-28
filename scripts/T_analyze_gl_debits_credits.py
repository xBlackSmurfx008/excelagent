#!/usr/bin/env python3
"""
Analyze GL Debits and Credits for Zero Balance Reconciliation
"""

import openpyxl

def main():
    # Load the enhanced Excel file
    file_path = 'data/05 May 2025 Reconciliation and Flex GL Activity.xlsx'
    wb = openpyxl.load_workbook(file_path)

    print('🧮 PROPER DEBIT/CREDIT CALCULATION FOR ZERO BALANCE')
    print('=' * 60)

    # Analyze all GL sheets for proper debit/credit totals
    gl_sheets = ['74400', '74505', '74510', '74515', '74520', '74525', '74530', '74535', '74540', '74550', '74560', '74570']

    total_debits = 0.0
    total_credits = 0.0
    gl_balances = {}

    print('\n📊 GL-BY-GL DEBIT/CREDIT ANALYSIS:')
    print('-' * 60)

    for gl_sheet_name in gl_sheets:
        if gl_sheet_name in wb.sheetnames:
            sheet = wb[gl_sheet_name]
            
            # Calculate debits and credits for this GL
            gl_debits = 0.0
            gl_credits = 0.0
            
            # Process all transaction rows
            for row in range(2, sheet.max_row + 1):  # Skip header
                # Column 8 = Debit, Column 9 = Credit
                debit_value = sheet.cell(row=row, column=8).value
                credit_value = sheet.cell(row=row, column=9).value
                
                if debit_value and isinstance(debit_value, (int, float)):
                    gl_debits += float(debit_value)
                
                if credit_value and isinstance(credit_value, (int, float)):
                    # Credits are already negative in the data, so we need to make them positive for proper accounting
                    gl_credits += abs(float(credit_value))
            
            # Calculate GL balance (Debits - Credits)
            gl_balance = gl_debits - gl_credits
            gl_balances[gl_sheet_name] = gl_balance
            
            # Add to totals
            total_debits += gl_debits
            total_credits += gl_credits
            
            print(f'GL {gl_sheet_name}:')
            print(f'   Debits: ${gl_debits:,.2f}')
            print(f'   Credits: ${gl_credits:,.2f}')
            print(f'   Balance: ${gl_balance:,.2f}')
            print()

    print('\n💰 TOTAL RECONCILIATION ANALYSIS:')
    print('-' * 60)
    print(f'Total Debits: ${total_debits:,.2f}')
    print(f'Total Credits: ${total_credits:,.2f}')
    print(f'Net Balance: ${total_debits - total_credits:,.2f}')
    print(f'Difference: ${abs(total_debits - total_credits):,.2f}')

    if abs(total_debits - total_credits) < 0.01:
        print('\n✅ RECONCILIATION BALANCED! (Debits = Credits)')
    else:
        print(f'\n❌ RECONCILIATION IMBALANCED! (Difference: ${abs(total_debits - total_credits):,.2f})')

    # Show GL balances
    print('\n📊 INDIVIDUAL GL BALANCES:')
    print('-' * 60)
    for gl, balance in gl_balances.items():
        status = 'BALANCED' if abs(balance) < 0.01 else 'IMBALANCED'
        print(f'GL {gl}: ${balance:,.2f} ({status})')

    # Show imbalanced GLs
    imbalanced_gls = [gl for gl, balance in gl_balances.items() if abs(balance) >= 0.01]
    if imbalanced_gls:
        print(f'\n⚠️ IMBALANCED GL ACCOUNTS: {len(imbalanced_gls)}')
        for gl in imbalanced_gls:
            print(f'   - GL {gl}: ${gl_balances[gl]:,.2f}')

    wb.close()

if __name__ == "__main__":
    main()
